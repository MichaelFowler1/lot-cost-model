"""The Excel output.

Most of these exist because of failures that produced a perfectly valid
workbook with something missing from the charts. openpyxl raises nothing when
you set an attribute a class does not have, so the only way to know a chart
feature survived is to read it back out of the XML.
"""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile

import openpyxl
import pytest

import lot_cost_model as M
import risk as R


@pytest.fixture(scope="module")
def plain_book(tmp_path_factory, analogy_df, estimate_df):
    """A workbook with no risk analysis."""
    path = tmp_path_factory.mktemp("wb") / "plain.xlsx"
    proj, ctx = M.run_lot_cost_model(analogy_df, estimate_df)
    M.save_complete_excel_workbook(
        str(path),
        proj,
        M.generate_analyst_summary(ctx, {"Program": "TEST"}),
        M.generate_fit_chart_data(ctx),
    )
    return path


@pytest.fixture(scope="module")
def risk_book(tmp_path_factory, analogy_df, estimate_df, cfg):
    """A workbook including the risk sheets, or skipped without cost_core."""
    if not R.AVAILABLE:
        pytest.skip(f"cost_core not installed: {R.IMPORT_ERROR}")
    path = tmp_path_factory.mktemp("wb") / "risk.xlsx"
    proj, ctx = M.run_lot_cost_model(analogy_df, estimate_df)
    res = R.run_risk(
        analogy_df["Qty"].to_numpy(float),
        analogy_df["AUC ($K)"].to_numpy(float),
        estimate_df["Qty"].to_numpy(int),
        estimate_df["Complexity"].to_numpy(float),
        cfg,
        R.RiskOptions(dollar_year=2025, n_iter=4000, seed=11),
    )
    M.save_complete_excel_workbook(
        str(path),
        proj,
        M.generate_analyst_summary(ctx, {"Program": "TEST"}),
        M.generate_fit_chart_data(ctx),
        R.summary_frame(res),
        res.intervals,
        res.scurve,
    )
    return path


def chart_xml(path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        names = sorted(
            n for n in z.namelist()
            if "charts/chart" in n and n.endswith(".xml")
        )
        return [z.read(n).decode("utf-8") for n in names]


# openpyxl serialises through lxml when it is installed and the standard
# library otherwise, and the two differ on trivia: <val v="1"/> against
# <val v="1" />. Assert on parsed elements so a missing lxml cannot fail a
# test about chart content.
def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def elements(xml: str, name: str, within: str | None = None):
    """Every element with this local tag name, optionally scoped to a parent."""
    root = ET.fromstring(xml)
    scopes = (
        [el for el in root.iter() if _local(el.tag) == within]
        if within
        else [root]
    )
    found = []
    for scope in scopes:
        found += [el for el in scope.iter() if _local(el.tag) == name]
    return found


def flag(xml: str, name: str, within: str | None = None) -> list[str | None]:
    """The `val` attribute of each matching element."""
    return [el.get("val") for el in elements(xml, name, within)]


class TestSheets:
    def test_three_sheets_without_risk(self, plain_book):
        assert openpyxl.load_workbook(plain_book).sheetnames == [
            "Analyst_Summary",
            "Estimate_Projections",
            "Fit_Chart_Data",
        ]

    def test_six_sheets_with_risk(self, risk_book):
        assert openpyxl.load_workbook(risk_book).sheetnames == [
            "Analyst_Summary",
            "Estimate_Projections",
            "Fit_Chart_Data",
            "Risk_Summary",
            "Risk_Intervals",
            "Risk_SCurve",
        ]


class TestChartsExist:
    def test_three_fit_charts(self, plain_book):
        wb = openpyxl.load_workbook(plain_book)
        assert len(wb["Fit_Chart_Data"]._charts) == 3

    def test_risk_sheets_each_carry_a_chart(self, risk_book):
        wb = openpyxl.load_workbook(risk_book)
        assert len(wb["Risk_Intervals"]._charts) == 1
        assert len(wb["Risk_SCurve"]._charts) == 1


class TestAxesAreVisible:
    """openpyxl writes delete="1" on a new axis, which hides it in Excel."""

    def test_no_axis_is_marked_deleted(self, plain_book):
        for xml in chart_xml(plain_book):
            deletes = flag(xml, "delete")
            assert deletes, "no delete flag written on either axis"
            assert "1" not in deletes
            assert len([d for d in deletes if d == "0"]) >= 2

    def test_tick_labels_are_positioned(self, plain_book):
        for xml in chart_xml(plain_book):
            assert "nextTo" in flag(xml, "tickLblPos")

    def test_titles_do_not_overlay_the_plot(self, plain_book):
        for xml in chart_xml(plain_book):
            assert "0" in flag(xml, "overlay")


class TestDataLabels:
    """A Series has no `dataLabels` alias, only `dLbls`. Assigning to the
    wrong one fails silently and draws nothing, which is what shipped."""

    def test_actual_auc_labels_reach_the_xml(self, plain_book):
        for xml in chart_xml(plain_book):
            assert elements(xml, "dLbls"), "no data labels written"
            assert flag(xml, "showVal", within="dLbls") == ["1"]

    def test_labels_do_not_also_print_the_series_name(self, plain_book):
        # Every show flag is written explicitly; Excel treats an absent flag
        # as inherited rather than false.
        xml = chart_xml(plain_book)[0]
        assert flag(xml, "showSerName", within="dLbls") == ["0"]
        assert flag(xml, "showLegendKey", within="dLbls") == ["0"]
        assert flag(xml, "showCatName", within="dLbls") == ["0"]

    def test_labels_are_shrunk_to_fit(self, plain_book):
        # 8pt, so the rate chart stays legible where lots of equal quantity
        # sit almost on top of each other.
        xml = chart_xml(plain_book)[0]
        sizes = [
            el.get("sz")
            for el in elements(xml, "defRPr", within="dLbls")
            if el.get("sz")
        ]
        assert "800" in sizes


class TestChartLayout:
    def test_anchors_are_spaced_wider_than_the_chart(self):
        # The three fit charts sit side by side. Their anchors are derived
        # from the chart width so widening one cannot overlap the next.
        width = 18
        cols = [
            openpyxl.utils.column_index_from_string(
                re.match(r"([A-Z]+)", M._chart_anchor(i, width)).group(1)
            )
            for i in range(3)
        ]
        step_cm = (cols[1] - cols[0]) * M._COL_CM
        assert cols == sorted(cols)
        assert step_cm > width, (
            f"charts step {step_cm:.1f}cm apart but are {width}cm wide"
        )

    def test_charts_sit_below_the_data(self, plain_book):
        wb = openpyxl.load_workbook(plain_book)
        ws = wb["Fit_Chart_Data"]
        for chart in ws._charts:
            assert chart.anchor._from.row + 1 > ws.max_row


class TestSCurveSheet:
    def test_one_row_per_percentile(self, risk_book):
        wb = openpyxl.load_workbook(risk_book)
        assert wb["Risk_SCurve"].max_row == 100  # header plus 99

    def test_named_markers_for_p50_and_p80(self, risk_book):
        ws = openpyxl.load_workbook(risk_book)["Risk_SCurve"]
        assert ws.cell(row=1, column=4).value.startswith("P50")
        assert ws.cell(row=1, column=6).value.startswith("P80")
        assert ws.cell(row=2, column=4).value == pytest.approx(0.50)
        assert ws.cell(row=2, column=6).value == pytest.approx(0.80)
        assert ws.cell(row=2, column=7).value > ws.cell(row=2, column=5).value

    def test_marker_names_carry_the_cost(self, risk_book):
        # The label has to name the percentile and show its cost, and the
        # series name is the only place both fit.
        ws = openpyxl.load_workbook(risk_book)["Risk_SCurve"]
        for name_col, cost_col in ((4, 5), (6, 7)):
            label = ws.cell(row=1, column=name_col).value
            cost = ws.cell(row=2, column=cost_col).value
            assert "$" in label
            assert label.split()[-1] == M._money_short(cost)

    def test_money_short_picks_a_sensible_unit(self):
        assert M._money_short(250_000_000) == "$250.0M"
        assert M._money_short(2_500_000_000) == "$2.5B"
        assert M._money_short(45_200) == "$45.2K"
        assert M._money_short(870) == "$870"

    def test_marker_labels_show_the_series_name(self, risk_book):
        xml = [x for x in chart_xml(risk_book) if "S-Curve" in x]
        assert xml, "no S-curve chart found"
        assert flag(xml[0], "showSerName", within="dLbls") == ["1", "1"]

    def test_curve_and_two_markers(self, risk_book):
        xml = [x for x in chart_xml(risk_book) if "S-Curve" in x][0]
        assert len(elements(xml, "ser")) == 3
