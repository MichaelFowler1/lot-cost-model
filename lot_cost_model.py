from __future__ import annotations

import json
import os
from datetime import datetime

import numpy as np
import openpyxl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.utils import get_column_letter
import pandas as pd

# ============================================================================
# 1. SETTINGS & CONFIGURATION
# ============================================================================
SETTINGS = {
    "AnalogyTableName": "AnalogyLots",
    "EstimateTableName": "EstimateLots",
    "CostUnitScale": 1.0,  # 1 = $K, 1000 = full dollars
    "TotalScale": 1000.0,  # Applied on top of CostUnitScale for totals
    # Reproduce a defect kept only for reconciling against legacy workbooks.
    # When True, the Rate model projects on the lot midpoint although it was
    # fitted against lot quantity, and the LC+Rate model drops its qty**c term
    # entirely. Both make projections that do not satisfy the equation the tool
    # prints, and because the rate exponent is negative the error is always
    # upward. Leave this False. See LEGACY_KEY below.
    "LegacyRateOmission": False,
    "DefaultCF": 1.0,
    "FitPriorUnits": 0,
    "FcstPriorUnits": 0,
    "SeedB": -0.152003093,
    "MaxIter": 100,
    "Tol": 1e-9,
    "RateSdFloor": 0.05,
    "SingularTol": 1e-12,
    "TGate": 2.0,
    "AiccTie": 2.0,
    "ToolVersion": None,  # filled from TOOL_VERSION; see provenance()
    "DefaultRunID": "R001",
    "DefaultProgram": "TEST",
    "DefaultRunLabel": "unlabeled run",
    "BaseYear": "",
}

#: The setting LegacyRateOmission replaced. Passing it is an error rather than
#: a no-op, because it used to default to the behaviour that is now off.
LEGACY_KEY = "ToolMatchProjection"

#: Bumped to 2.1.0 for the rate-projection correction. Anything a 2.0 build
#: produced with Rate or LC+Rate selected is overstated, and the version was
#: hardcoded to "2.0-dev" for every run, so an old workbook cannot be dated
#: from the inside. Recording it properly is the point of provenance() below.
TOOL_VERSION = "2.1.0"


def _source_revision() -> str:
    """Short git revision of this checkout, when it is one.

    An installed copy or a downloaded zip has no repository, which is not an
    error: the version string still identifies the release.
    """
    import subprocess

    try:
        here = os.path.dirname(os.path.abspath(__file__))
        out = subprocess.run(
            ["git", "-C", here, "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if out.returncode == 0 and out.stdout.strip():
            dirty = subprocess.run(
                ["git", "-C", here, "status", "--porcelain"],
                capture_output=True,
                text=True,
                timeout=5,
            )
            suffix = "+modified" if dirty.stdout.strip() else ""
            return out.stdout.strip() + suffix
    except Exception:
        pass
    return ""


def provenance(cfg: dict | None = None) -> dict:
    """What produced this run, so a saved workbook can be identified later.

    A workbook that cannot say which build made it also cannot say whether it
    predates a correction, which is exactly the position every estimate this
    tool produced before 2.1.0 is in.
    """
    cfg = cfg or SETTINGS
    rev = _source_revision()
    legacy = bool(cfg.get("LegacyRateOmission", False))
    return {
        "Tool version": TOOL_VERSION + (f" ({rev})" if rev else ""),
        "Run timestamp": datetime.now().astimezone().strftime(
            "%Y-%m-%d %H:%M:%S %Z"
        ),
        "Rate projection": (
            "LEGACY - Rate and LC+Rate are overstated; for reconciling a "
            "pre-2.1.0 workbook only"
            if legacy
            else "corrected (projections satisfy the fitted equation)"
        ),
    }


# ============================================================================
# 2. SHARED MATHEMATICAL & STATISTICAL HELPERS
# ============================================================================
def find_col(df_columns: list, candidates: list) -> str | None:
    """Case-insensitive search for the first matching column name."""
    col_map = {c.strip().lower(): c for c in df_columns}
    for cand in candidates:
        if cand.strip().lower() in col_map:
            return col_map[cand.strip().lower()]
    return None


def to_num(val):
    """Clean string currency/commas and convert to numeric float."""
    if pd.isna(val):
        return np.nan
    if isinstance(val, (int, float, np.number)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.replace("$", "").replace(",", "").strip()
        try:
            return float(cleaned) if cleaned != "" else np.nan
        except ValueError:
            return np.nan
    return np.nan


def lmp_func(
    s: float, e: float, q: float, b: float | None
) -> float | None:
    """Calculate Lot Midpoint (LMP) given Start, End, Qty, and b-slope."""
    if b is None or pd.isna(b):
        return np.nan
    if q <= 1:
        return float(s)
    if abs(b) < 1e-12:
        return (s + e) / 2.0
    if abs(b + 1.0) < 1e-6:
        lo = max(s - 0.5, 1e-6)
        return q / (np.log(e + 0.5) - np.log(lo))

    p = b + 1.0
    lo = max(s - 0.5, 1e-6)
    v = ((e + 0.5) ** p - lo**p) / (p * q)
    if v <= 0:
        return np.nan
    return v ** (1.0 / b)


def ols_fit(
    x_cols: list[np.ndarray], y: np.ndarray, singular_tol: float = 1e-12
):
    """Ordinary Least Squares matching Excel/M Fit Space Statistics."""
    n = len(y)
    k = len(x_cols) + 1  # Intercept + predictors
    x = np.column_stack([np.ones(n)] + x_cols)

    xtx = x.T @ x
    xty = x.T @ y

    det = np.linalg.det(xtx)
    diag_prod = np.prod(np.abs(np.diag(xtx)))
    if diag_prod <= 0 or (abs(det) / diag_prod) < singular_tol:
        return None

    try:
        inv = np.linalg.inv(xtx)
    except np.linalg.LinAlgError:
        return None

    beta = inv @ xty
    fitted = x @ beta
    res = y - fitted
    sse = float(np.sum(res**2))
    ybar = float(np.mean(y))
    sst = float(np.sum((y - ybar) ** 2))
    ssr = sst - sse
    df = n - k

    s2 = (sse / df) if df > 0 else np.nan
    r2 = (1.0 - (sse / sst)) if sst > 0 else np.nan
    ar2 = (
        (1.0 - (1.0 - r2) * (n - 1) / df)
        if (df > 0 and pd.notna(r2))
        else np.nan
    )
    fstat = (
        ((ssr / (k - 1)) / (sse / df))
        if (df > 0 and k > 1 and sse > 0)
        else np.nan
    )

    se = []
    for a in range(k):
        v = s2 * inv[a, a] if pd.notna(s2) else np.nan
        se.append(np.sqrt(v) if pd.notna(v) and v > 0 else np.nan)

    return {
        "Beta": beta.tolist(),
        "SE": se,
        "Fitted": fitted,
        "SSE": sse,
        "InvDiag": np.diag(inv).tolist(),
        "R2": r2,
        "AdjR2": ar2,
        "SEy": np.sqrt(s2) if pd.notna(s2) and s2 > 0 else np.nan,
        "F": fstat,
        "DF": df,
        "SSreg": ssr,
        "SSresid": sse,
        "N": n,
        "K": k,
    }


def solve_model(
    fit_q: np.ndarray,
    fit_c: np.ndarray,
    fit_se: list[dict],
    use_rate: bool,
    cfg: dict,
):
    """Iterative solver for Learning Curve parameter b (Goal Seek equivalent)."""
    ln_y = np.log(fit_c)
    ln_r = np.log(fit_q)

    def mid_at(b_val):
        return np.array(
            [
                np.log(lmp_func(se["S"], se["E"], q, b_val))
                for se, q in zip(fit_se, fit_q)
            ]
        )

    b = cfg["SeedB"]
    delta = 1.0
    iteration = 0

    while iteration < cfg["MaxIter"] and delta > cfg["Tol"]:
        bp = b
        x_pred = [mid_at(bp), ln_r] if use_rate else [mid_at(bp)]
        fit = ols_fit(x_pred, ln_y, cfg["SingularTol"])
        if fit is None:
            return None
        b = fit["Beta"][1]
        delta = abs(b - bp)
        iteration += 1

    x_pred = [mid_at(b), ln_r] if use_rate else [mid_at(b)]
    final_fit = ols_fit(x_pred, ln_y, cfg["SingularTol"])
    if final_fit is None:
        return None

    resid = abs(final_fit["Beta"][1] - b)
    final_fit["Iter"] = iteration
    final_fit["Delta"] = resid
    final_fit["Converged"] = resid <= cfg["Tol"]
    return final_fit


def track_units(quantities: np.ndarray, prior: int):
    cums = np.cumsum(quantities) + prior
    starts = cums - quantities + 1
    return [{"S": s, "E": e} for s, e in zip(starts, cums)]


# ============================================================================
# 3. QUERY 1: CORE PROJECTIONS FACT TABLE
# ============================================================================
def run_lot_cost_model(
    analogy_df: pd.DataFrame,
    estimate_df: pd.DataFrame,
    config_overrides: dict | None = None,
) -> tuple[pd.DataFrame, dict]:
    cfg = SETTINGS.copy()
    if config_overrides:
        if LEGACY_KEY in config_overrides:
            # Silently honouring the new default would hand a caller who asked
            # for legacy behaviour a different set of numbers without saying so.
            raise ValueError(
                f"'{LEGACY_KEY}' was replaced by 'LegacyRateOmission'.\n\n"
                f"{LEGACY_KEY}=True is now LegacyRateOmission=True, and "
                f"{LEGACY_KEY}=False is now LegacyRateOmission=False. The "
                "default changed: projections now satisfy the fitted equation, "
                "which lowers any estimate where Rate or LC+Rate was selected."
            )
        cfg.update(config_overrides)

    if analogy_df.empty:
        raise ValueError(
            f"Analogy table '{cfg['AnalogyTableName']}' contains no rows."
        )
    if estimate_df.empty:
        raise ValueError(
            f"Estimate table '{cfg['EstimateTableName']}' contains no rows."
        )

    # Resolve Columns - Analogy Table
    a_cols = list(analogy_df.columns)
    a_yr = find_col(
        a_cols,
        ["Year", "FY", "Fiscal Year", "Lot FY", "Lot Year", "FY of Lot"],
    )
    a_q = find_col(
        a_cols,
        [
            "Analogy Qtys",
            "Analogy Qty",
            "AnalogyQtys",
            "AnalogyQty",
            "Analogy Quantity",
            "Qty",
            "Qtys",
            "Quantity",
            "Lot Qty",
            "Lot Quantity",
        ],
    )
    a_c = find_col(
        a_cols,
        [
            "AnalogyUnitCost_K",
            "AnalogyUnitCost_Dollars",
            "AnalogyUnitCost",
            "Analogy Unit Cost",
            "AUC",
            "CP24",
            "AUC_K",
            "AUC ($K)",
            "Unit Cost",
            "UnitCost",
        ],
    )
    a_seq = find_col(
        a_cols,
        [
            "AnalogySeq",
            "Analogy Seq",
            "Analogy Lot",
            "Seq",
            "Sequence",
            "Lot",
            "Lot No",
            "Lot No.",
            "LotNo",
        ],
    )

    if not a_q or not a_c:
        raise ValueError(
            f"Analogy table needs a quantity column and a unit-cost column. Seen: {a_cols}"
        )

    df_a = analogy_df.copy()
    df_a["RowNo"] = np.arange(len(df_a))
    df_a["AQty"] = df_a[a_q].apply(to_num)
    df_a["ACost"] = df_a[a_c].apply(to_num)
    df_a["_Seq"] = (
        df_a[a_seq].apply(to_num) if a_seq else np.nan
    )
    df_a["_Yr"] = df_a[a_yr].apply(to_num) if a_yr else np.nan

    unit_keep = df_a[
        df_a["AQty"].notna() & (df_a["AQty"] > 0)
    ].copy()
    n_unit = len(unit_keep)
    seq_ok = (
        n_unit > 0
        and unit_keep["_Seq"].notna().sum() == n_unit
    )
    yr_ok = (
        n_unit > 0
        and unit_keep["_Yr"].notna().sum() == n_unit
    )

    if seq_ok:
        unit_set = unit_keep.sort_values(
            by=["_Seq", "RowNo"], ascending=[True, True]
        ).reset_index(drop=True)
    elif yr_ok:
        unit_set = unit_keep.sort_values(
            by=["_Yr", "RowNo"], ascending=[True, True]
        ).reset_index(drop=True)
    else:
        unit_set = unit_keep.sort_values(
            by=["RowNo"], ascending=[True]
        ).reset_index(drop=True)

    fit_set = unit_set[
        unit_set["ACost"].notna() & (unit_set["ACost"] > 0)
    ].reset_index(drop=True)
    n_keep = len(fit_set)

    # Resolve Columns - Estimate Table
    e_cols = list(estimate_df.columns)
    c_lot = find_col(
        e_cols,
        [
            "Lot",
            "Lot #",
            "LRIP",
            "Lot ID",
            "Lot No",
            "Lot No.",
            "LotNo",
            "LRIP Lot",
        ],
    )
    c_yr = find_col(
        e_cols,
        [
            "Year",
            "FY",
            "Fiscal Year",
            "Lot FY",
            "Lot Year",
            "Delivery FY",
            "Buy FY",
        ],
    )
    c_qty = find_col(
        e_cols,
        [
            "Qty",
            "Quantity",
            "Units",
            "Lot Qty",
            "Lot Quantity",
            "Estimate Qty",
            "Buy Qty",
            "LRIP Qty",
        ],
    )
    c_cf = find_col(
        e_cols,
        [
            "ComplexityFactor",
            "Complexity Factor",
            "Complexity/CP",
            "Complexity",
            "CF",
        ],
    )

    if not c_qty:
        raise ValueError(
            f"Estimate table needs a quantity column. Seen: {e_cols}"
        )

    df_e = estimate_df.copy()
    df_e["RowNo"] = np.arange(len(df_e))
    df_e["Lot"] = (
        df_e[c_lot].fillna("").astype(str)
        if c_lot
        else [f"Lot {i+1}" for i in range(len(df_e))]
    )
    df_e["Year"] = (
        df_e[c_yr].apply(to_num) if c_yr else np.nan
    )
    df_e["Qty"] = df_e[c_qty].apply(to_num)
    df_e["ComplexityFactor"] = (
        df_e[c_cf].apply(to_num) if c_cf else np.nan
    )

    fcst_ord = (
        df_e[df_e["Qty"].notna() & (df_e["Qty"] > 0)]
        .sort_values(
            by=["Year", "RowNo"], ascending=[True, True]
        )
        .reset_index(drop=True)
    )

    cf_vals = []
    last_cf = cfg["DefaultCF"]
    for val in fcst_ord["ComplexityFactor"]:
        if pd.isna(val) or val <= 0:
            cf_vals.append(last_cf)
        else:
            last_cf = float(val)
            cf_vals.append(last_cf)
    fcst_ord["ComplexityFactor"] = cf_vals

    if n_keep < 3:
        raise ValueError(
            f"Learning curve needs at least 3 analogy lots with both quantity and cost. Found: {n_keep}"
        )
    if len(fcst_ord) < 1:
        raise ValueError("No forecast rows found.")

    # Unit Tracking
    unit_q = unit_set["AQty"].to_numpy()
    unit_se = track_units(unit_q, cfg["FitPriorUnits"])
    complete_idx = unit_set[
        unit_set["ACost"].notna() & (unit_set["ACost"] > 0)
    ].index.tolist()

    fit_q = fit_set["AQty"].to_numpy()
    fit_c = fit_set["ACost"].to_numpy()
    fit_se = [unit_se[i] for i in complete_idx]

    fcst_q = fcst_ord["Qty"].to_numpy()
    fcst_se = track_units(fcst_q, cfg["FcstPriorUnits"])

    # Model Fitting
    ln_r = np.log(fit_q)
    ln_y = np.log(fit_c)
    rate_sd = (
        np.std(ln_r, ddof=1) if len(ln_r) > 1 else np.nan
    )

    rate_why = ""
    if n_keep < 4:
        rate_why = "fewer than 4 analogy lots"
    elif pd.isna(rate_sd):
        rate_why = "no spread in ln(lot qty)"
    elif rate_sd < cfg["RateSdFloor"]:
        rate_why = (
            "lot quantities too uniform for a rate term"
        )

    rate_ok = rate_why == ""

    mdl_lc = solve_model(
        fit_q, fit_c, fit_se, use_rate=False, cfg=cfg
    )
    mdl_rt = (
        ols_fit([ln_r], ln_y, cfg["SingularTol"])
        if rate_ok
        else None
    )
    mdl_lcr = (
        solve_model(
            fit_q, fit_c, fit_se, use_rate=True, cfg=cfg
        )
        if rate_ok
        else None
    )

    if mdl_lc is None:
        raise RuntimeError("Learning curve fit failed.")

    gb = (
        lambda m, i: (
            m["Beta"][i]
            if (m and len(m["Beta"]) > i)
            else np.nan
        )
    )
    gs = (
        lambda m, i: (
            m["SE"][i] if (m and len(m["SE"]) > i) else np.nan
        )
    )
    gf = lambda m, f: m.get(f, np.nan) if m else np.nan
    slope = (
        lambda b_val: (
            round((2**b_val) * 100, 2)
            if pd.notna(b_val)
            else np.nan
        )
    )
    t1_of = (
        lambda m: (
            round(
                np.exp(m["Beta"][0]) * cfg["CostUnitScale"], 2
            )
            if (m and "Beta" in m)
            else np.nan
        )
    )

    def stat_msg(m, name):
        if m is None:
            return f"{name} suppressed: {rate_why}"
        if not m.get("Converged", True):
            return f"{name} NOT CONVERGED"
        return f"{name} ok"

    fit_status = f"{stat_msg(mdl_lc, 'LC')}; {stat_msg(mdl_rt, 'Rate')}; {stat_msg(mdl_lcr, 'LC+Rate')}"
    data_source = f"{cfg['AnalogyTableName']} + {cfg['EstimateTableName']}"

    t1_lc, b_lc = np.exp(mdl_lc["Beta"][0]), gb(mdl_lc, 1)
    t1_rt, b_rt = (
        (np.exp(mdl_rt["Beta"][0]), gb(mdl_rt, 1))
        if mdl_rt
        else (np.nan, np.nan)
    )
    t1_br, b_br, c_br = (
        (
            np.exp(mdl_lcr["Beta"][0]),
            gb(mdl_lcr, 1),
            gb(mdl_lcr, 2),
        )
        if mdl_lcr
        else (np.nan, np.nan, np.nan)
    )

    # Projections on Forecast Lots
    res_df = fcst_ord.copy()
    res_df["FirstUnitInLot"] = [se["S"] for se in fcst_se]
    res_df["LastUnitInLot"] = [se["E"] for se in fcst_se]

    res_df["LC_LMP"] = [
        lmp_func(s, e, q, b_lc)
        for s, e, q in zip(
            res_df["FirstUnitInLot"],
            res_df["LastUnitInLot"],
            res_df["Qty"],
        )
    ]
    res_df["LC_UnitCost"] = (
        t1_lc
        * (res_df["LC_LMP"] ** b_lc)
        * cfg["CostUnitScale"]
    )
    res_df["LC_BaseTotal"] = (
        res_df["LC_UnitCost"]
        * res_df["Qty"]
        * cfg["TotalScale"]
    )
    res_df["LC_AdjTotal"] = (
        res_df["LC_BaseTotal"] * res_df["ComplexityFactor"]
    )

    res_df["RT_LMP"] = [
        lmp_func(s, e, q, b_rt)
        for s, e, q in zip(
            res_df["FirstUnitInLot"],
            res_df["LastUnitInLot"],
            res_df["Qty"],
        )
    ]
    if pd.notna(t1_rt):
        if cfg["LegacyRateOmission"]:
            # Wrong on purpose: the Rate model was fitted against lot
            # quantity, so projecting on the midpoint evaluates it at a
            # different variable than it was fitted on.
            res_df["RT_UnitCost"] = (
                t1_rt
                * (res_df["RT_LMP"] ** b_rt)
                * cfg["CostUnitScale"]
            )
        else:
            res_df["RT_UnitCost"] = (
                t1_rt
                * (res_df["Qty"] ** b_rt)
                * cfg["CostUnitScale"]
            )
        res_df["RT_BaseTotal"] = (
            res_df["RT_UnitCost"]
            * res_df["Qty"]
            * cfg["TotalScale"]
        )
        res_df["RT_AdjTotal"] = (
            res_df["RT_BaseTotal"] * res_df["ComplexityFactor"]
        )
    else:
        res_df["RT_UnitCost"] = res_df["RT_BaseTotal"] = (
            res_df["RT_AdjTotal"]
        ) = np.nan

    res_df["LCR_LMP"] = [
        lmp_func(s, e, q, b_br)
        for s, e, q in zip(
            res_df["FirstUnitInLot"],
            res_df["LastUnitInLot"],
            res_df["Qty"],
        )
    ]
    if pd.notna(t1_br):
        # Dropping qty**c evaluates the fit at a lot quantity of one unit
        # while keeping the real lot's learning position, which is not a
        # production rate anybody chose to hold it at.
        rate_factor = (
            1.0
            if cfg["LegacyRateOmission"]
            else (res_df["Qty"] ** c_br)
        )
        res_df["LCR_UnitCost"] = (
            t1_br
            * (res_df["LCR_LMP"] ** b_br)
            * rate_factor
            * cfg["CostUnitScale"]
        )
        res_df["LCR_BaseTotal"] = (
            res_df["LCR_UnitCost"]
            * res_df["Qty"]
            * cfg["TotalScale"]
        )
        res_df["LCR_AdjTotal"] = (
            res_df["LCR_BaseTotal"]
            * res_df["ComplexityFactor"]
        )
    else:
        res_df["LCR_UnitCost"] = res_df["LCR_BaseTotal"] = (
            res_df["LCR_AdjTotal"]
        ) = np.nan

    # Fit Statistics Columns
    rnd = (
        lambda v, d: (
            round(v, d) if pd.notna(v) else np.nan
        )
    )
    pct = (
        lambda v: (
            round(v * 100, 2) if pd.notna(v) else np.nan
        )
    )

    res_df["LC_T1"] = t1_of(mdl_lc)
    res_df["LC_Icept"] = rnd(gb(mdl_lc, 0), 4)
    res_df["LC_IceptSE"] = rnd(gs(mdl_lc, 0), 4)
    res_df["LC_Learn"] = rnd(gb(mdl_lc, 1), 4)
    res_df["LC_LearnSE"] = rnd(gs(mdl_lc, 1), 4)
    res_df["LC_LearnSlope"] = slope(gb(mdl_lc, 1))
    res_df["LC_R2"] = pct(gf(mdl_lc, "R2"))
    res_df["LC_SEy"] = rnd(gf(mdl_lc, "SEy"), 4)
    res_df["LC_F"] = rnd(gf(mdl_lc, "F"), 2)
    res_df["LC_df"] = gf(mdl_lc, "DF")
    res_df["LC_SSreg"] = rnd(gf(mdl_lc, "SSreg"), 4)
    res_df["LC_SSresid"] = rnd(gf(mdl_lc, "SSresid"), 4)
    res_df["LC_AdjR2"] = pct(gf(mdl_lc, "AdjR2"))

    res_df["RT_T1"] = t1_of(mdl_rt)
    res_df["RT_Icept"] = rnd(gb(mdl_rt, 0), 4)
    res_df["RT_IceptSE"] = rnd(gs(mdl_rt, 0), 4)
    res_df["RT_Rate"] = rnd(gb(mdl_rt, 1), 4)
    res_df["RT_RateSE"] = rnd(gs(mdl_rt, 1), 4)
    res_df["RT_RateSlope"] = slope(gb(mdl_rt, 1))
    res_df["RT_R2"] = pct(gf(mdl_rt, "R2"))
    res_df["RT_SEy"] = rnd(gf(mdl_rt, "SEy"), 4)
    res_df["RT_F"] = rnd(gf(mdl_rt, "F"), 2)
    res_df["RT_df"] = gf(mdl_rt, "DF")
    res_df["RT_SSreg"] = rnd(gf(mdl_rt, "SSreg"), 4)
    res_df["RT_SSresid"] = rnd(gf(mdl_rt, "SSresid"), 4)
    res_df["RT_AdjR2"] = pct(gf(mdl_rt, "AdjR2"))

    res_df["LCR_T1"] = t1_of(mdl_lcr)
    res_df["LCR_Icept"] = rnd(gb(mdl_lcr, 0), 4)
    res_df["LCR_IceptSE"] = rnd(gs(mdl_lcr, 0), 4)
    res_df["LCR_Learn"] = rnd(gb(mdl_lcr, 1), 4)
    res_df["LCR_LearnSE"] = rnd(gs(mdl_lcr, 1), 4)
    res_df["LCR_LearnSlope"] = slope(gb(mdl_lcr, 1))
    res_df["LCR_Rate"] = rnd(gb(mdl_lcr, 2), 4)
    res_df["LCR_RateSE"] = rnd(gs(mdl_lcr, 2), 4)
    res_df["LCR_RateSlope"] = slope(gb(mdl_lcr, 2))
    res_df["LCR_R2"] = pct(gf(mdl_lcr, "R2"))
    res_df["LCR_SEy"] = rnd(gf(mdl_lcr, "SEy"), 4)
    res_df["LCR_F"] = rnd(gf(mdl_lcr, "F"), 2)
    res_df["LCR_df"] = gf(mdl_lcr, "DF")
    res_df["LCR_SSreg"] = rnd(gf(mdl_lcr, "SSreg"), 4)
    res_df["LCR_SSresid"] = rnd(gf(mdl_lcr, "SSresid"), 4)
    res_df["LCR_AdjR2"] = pct(gf(mdl_lcr, "AdjR2"))

    res_df["FitStatus"] = fit_status
    res_df["DataSource"] = data_source

    if "Lot" not in res_df.columns:
        res_df["Lot"] = [
            f"Lot {i+1}" for i in range(len(res_df))
        ]
    if "Year" not in res_df.columns:
        res_df["Year"] = np.nan
    if "ComplexityFactor" not in res_df.columns:
        res_df["ComplexityFactor"] = cfg["DefaultCF"]

    round_cols_4 = [
        "LC_LMP",
        "RT_LMP",
        "LCR_LMP",
        "ComplexityFactor",
    ]
    round_cols_2 = [
        "LC_UnitCost",
        "LC_BaseTotal",
        "LC_AdjTotal",
        "RT_UnitCost",
        "RT_BaseTotal",
        "RT_AdjTotal",
        "LCR_UnitCost",
        "LCR_BaseTotal",
        "LCR_AdjTotal",
    ]
    for c in round_cols_4:
        res_df[c] = res_df[c].apply(lambda v: rnd(v, 4))
    for c in round_cols_2:
        res_df[c] = res_df[c].apply(lambda v: rnd(v, 2))

    res_df["Lot"] = res_df["Lot"].astype(str)
    res_df["Year"] = pd.to_numeric(
        res_df["Year"], errors="coerce"
    ).astype("Int64")
    res_df["Qty"] = pd.to_numeric(
        res_df["Qty"], errors="coerce"
    ).astype("Int64")
    res_df["FirstUnitInLot"] = pd.to_numeric(
        res_df["FirstUnitInLot"], errors="coerce"
    ).astype("Int64")
    res_df["LastUnitInLot"] = pd.to_numeric(
        res_df["LastUnitInLot"], errors="coerce"
    ).astype("Int64")

    rename_dict = {
        "Lot": "Lot",
        "Year": "Fiscal Year",
        "Qty": "Lot Quantity",
        "FirstUnitInLot": "First Unit in Lot",
        "LastUnitInLot": "Last Unit in Lot",
        "ComplexityFactor": "Complexity Factor",
        "LC_LMP": "LC Lot Midpoint (unit no.)",
        "LC_UnitCost": "LC Unit Cost ($K)",
        "LC_BaseTotal": "LC Lot Cost Before Complexity ($)",
        "LC_AdjTotal": "LC Lot Cost After Complexity ($)",
        "RT_LMP": "Rate Lot Midpoint (unit no.)",
        "RT_UnitCost": "Rate Unit Cost ($K)",
        "RT_BaseTotal": "Rate Lot Cost Before Complexity ($)",
        "RT_AdjTotal": "Rate Lot Cost After Complexity ($)",
        "LCR_LMP": "LC+Rate Lot Midpoint (unit no.)",
        "LCR_UnitCost": "LC+Rate Unit Cost ($K)",
        "LCR_BaseTotal": (
            "LC+Rate Lot Cost Before Complexity ($)"
        ),
        "LCR_AdjTotal": (
            "LC+Rate Lot Cost After Complexity ($)"
        ),
        "LC_T1": "LC T1 First Unit Cost ($K)",
        "LC_Icept": "LC Intercept Coeff",
        "LC_IceptSE": "LC Intercept SE",
        "LC_Learn": "LC Learning Coeff",
        "LC_LearnSE": "LC Learning SE",
        "LC_LearnSlope": "LC Learning Slope (%)",
        "LC_R2": "LC R2 (%)",
        "LC_SEy": "LC SEy",
        "LC_F": "LC F",
        "LC_df": "LC df",
        "LC_SSreg": "LC SSreg",
        "LC_SSresid": "LC SSresid",
        "LC_AdjR2": "LC Adj R2 (%)",
        "RT_T1": "Rate T1 First Unit Cost ($K)",
        "RT_Icept": "Rate Intercept Coeff",
        "RT_IceptSE": "Rate Intercept SE",
        "RT_Rate": "Rate Coeff",
        "RT_RateSE": "Rate SE",
        "RT_RateSlope": "Rate Slope (%)",
        "RT_R2": "Rate R2 (%)",
        "RT_SEy": "Rate SEy",
        "RT_F": "Rate F",
        "RT_df": "Rate df",
        "RT_SSreg": "Rate SSreg",
        "RT_SSresid": "Rate SSresid",
        "RT_AdjR2": "Rate Adj R2 (%)",
        "LCR_T1": "LC+Rate T1 First Unit Cost ($K)",
        "LCR_Icept": "LC+Rate Intercept Coeff",
        "LCR_IceptSE": "LC+Rate Intercept SE",
        "LCR_Learn": "LC+Rate Learning Coeff",
        "LCR_LearnSE": "LC+Rate Learning SE",
        "LCR_LearnSlope": "LC+Rate Learning Slope (%)",
        "LCR_Rate": "LC+Rate Rate Coeff",
        "LCR_RateSE": "LC+Rate Rate SE",
        "LCR_RateSlope": "LC+Rate Rate Slope (%)",
        "LCR_R2": "LC+Rate R2 (%)",
        "LCR_SEy": "LC+Rate SEy",
        "LCR_F": "LC+Rate F",
        "LCR_df": "LC+Rate df",
        "LCR_SSreg": "LC+Rate SSreg",
        "LCR_SSresid": "LC+Rate SSresid",
        "LCR_AdjR2": "LC+Rate Adj R2 (%)",
        "FitStatus": "Fit Status",
        "DataSource": "Input Table",
    }

    ordered_cols = list(rename_dict.keys())
    projections_df = res_df[ordered_cols].rename(
        columns=rename_dict
    )

    models_context = {
        "mdl_lc": mdl_lc,
        "mdl_rt": mdl_rt,
        "mdl_lcr": mdl_lcr,
        "fit_q": fit_q,
        "fit_c": fit_c,
        "fit_se": fit_se,
        "n_keep": n_keep,
        "n_unit": n_unit,
        "rate_sd": rate_sd,
        "rate_ok": rate_ok,
        "rate_why": rate_why,
        "t1_lc": t1_lc,
        "b_lc": b_lc,
        "t1_rt": t1_rt,
        "b_rt": b_rt,
        "t1_br": t1_br,
        "b_br": b_br,
        "c_br": c_br,
        "cfg": cfg,
    }

    return projections_df, models_context


# ============================================================================
# 4. QUERY 2: ANALYST SUMMARY ENGINE (Human-Readable Report)
# ============================================================================
def generate_analyst_summary(
    models_context: dict, run_info: dict | None = None
) -> pd.DataFrame:
    ctx = models_context
    cfg = ctx["cfg"]
    ri = run_info or {}

    run_id = ri.get("RunID", cfg["DefaultRunID"])
    program = ri.get("Program", cfg["DefaultProgram"])
    run_label = ri.get("RunLabel", cfg["DefaultRunLabel"])
    base_year = ri.get("BaseYear", cfg["BaseYear"])

    cost_basis_txt = (
        "NOT STATED - declare BaseYear in the RunInfo table"
        if not base_year
        else f"BY{base_year} $K as entered (no escalation applied by the tool)"
    )

    n_keep = ctx["n_keep"]
    n_unit = ctx["n_unit"]
    fit_c = ctx["fit_c"]
    rate_sd = ctx["rate_sd"]
    rate_ok = ctx["rate_ok"]
    rate_why = ctx["rate_why"]

    ln_y = np.log(fit_c)
    ybar = np.mean(ln_y)
    sst = np.sum((ln_y - ybar) ** 2)

    def stat_for(m, k, rate_idx=None):
        if m is None:
            return None
        sse0 = m["SSE"]
        dfe = n_keep - k
        see = np.sqrt(sse0 / dfe) if dfe > 0 else None
        r2 = (1.0 - sse0 / sst) if sst > 0 else None
        adj = (
            (1.0 - (1.0 - r2) * (n_keep - 1) / dfe)
            if (r2 is not None and dfe > 0)
            else None
        )
        cv = (
            np.sqrt(np.exp(see * see) - 1.0)
            if see is not None
            else None
        )

        fit_u = np.exp(m["Fitted"])
        mape = np.mean(np.abs(fit_c - fit_u) / fit_c)
        bias = np.mean(fit_c / fit_u - 1.0)

        kp = k + 1
        sseg = max(sse0, 1e-30)
        aicc = (
            (
                n_keep * np.log(sseg / n_keep)
                + 2 * kp
                + 2 * kp * (kp + 1) / (n_keep - kp - 1)
            )
            if (n_keep - kp - 1 > 0)
            else None
        )

        sec = (
            see * np.sqrt(m["InvDiag"][rate_idx])
            if (rate_idx is not None and see is not None)
            else None
        )
        tc = (
            (m["Beta"][rate_idx] / sec)
            if (sec is not None and sec >= 1e-15)
            else None
        )

        return {
            "SEE": see,
            "R2": r2,
            "Adj": adj,
            "CV": cv,
            "MAPE": mape,
            "Bias": bias,
            "AICc": aicc,
            "T": tc,
        }

    s_lc = stat_for(ctx["mdl_lc"], 2, None)
    s_rt = stat_for(ctx["mdl_rt"], 2, 1)
    s_lcr = stat_for(ctx["mdl_lcr"], 3, 2)

    gs = lambda s, f: s.get(f) if s else None

    t_gate = cfg["TGate"]
    aicc_tie = cfg["AiccTie"]

    lcr_gate = (
        s_lcr is not None
        and gs(s_lcr, "T") is not None
        and abs(gs(s_lcr, "T")) >= t_gate
    )
    rt_gate = (
        s_rt is not None
        and gs(s_rt, "T") is not None
        and abs(gs(s_rt, "T")) >= t_gate
    )

    if lcr_gate:
        sel = "LC+Rate"
    elif (
        ctx["mdl_lc"] is None and ctx["mdl_rt"] is not None
    ):
        sel = "Rate"
    elif (
        ctx["mdl_lc"] is not None
        and rt_gate
        and gs(s_rt, "AICc") is not None
        and gs(s_lc, "AICc") is not None
        and (
            gs(s_rt, "AICc") + aicc_tie < gs(s_lc, "AICc")
        )
    ):
        sel = "Rate"
    elif ctx["mdl_lc"] is not None:
        sel = "LC"
    else:
        sel = None

    aicc_vals = [
        v
        for v in [
            gs(s_lc, "AICc"),
            gs(s_rt, "AICc"),
            gs(s_lcr, "AICc"),
        ]
        if v is not None
    ]
    best_aicc = min(aicc_vals) if aicc_vals else None

    def d_aicc(s):
        if (
            s is None
            or gs(s, "AICc") is None
            or best_aicc is None
        ):
            return None
        return gs(s, "AICc") - best_aicc

    lcr_aicc_disagrees = (
        sel == "LC+Rate"
        and gs(s_lcr, "AICc") is not None
        and gs(s_lc, "AICc") is not None
        and (
            gs(s_lc, "AICc") + aicc_tie
            < gs(s_lcr, "AICc")
        )
    )

    rate_why_not = (
        ""
        if rate_ok
        else (
            "needs >= 4 costed lots"
            if n_keep < 4
            else f"SD(ln qty) {rate_sd:.4f} < {cfg['RateSdFloor']} floor"
        )
    )

    if sel == "LC+Rate":
        sel_note = (
            f"Rate coefficient significant (|t| >= {t_gate})."
            + (
                " Note: AICc favors LC at this sample size - state both in the BOE."
                if lcr_aicc_disagrees
                else ""
            )
        )
    elif sel == "Rate":
        sel_note = f"Slope significant and beats LC by more than {aicc_tie} AICc."
    elif sel == "LC":
        base_msg = "Default model. "
        if not rate_ok:
            base_msg += (
                f"Rate models gated off ({rate_why_not})."
            )
        elif (
            ctx["mdl_lcr"] is not None and not lcr_gate
        ):
            base_msg += f"Rate coefficient not significant (|t| < {t_gate})."
        if gs(s_lc, "AICc") is None:
            base_msg += " n too small for AICc; comparison on coefficient significance only."
        sel_note = base_msg
    else:
        sel_note = "No model could be fitted."

    def fit_txt(m):
        if m is not None:
            return "Yes"
        if not rate_ok:
            return f"No - rate gate ({rate_why_not})"
        return "No - did not converge or singular fit"

    fmt_n = (
        lambda v, f: (
            "n/a"
            if (v is None or pd.isna(v))
            else (
                f"{v:,.2f}"
                if f == "#,##0.00"
                else (
                    f"{v:.4f}"
                    if f == "0.0000"
                    else (
                        f"{v:.6f}"
                        if f == "0.000000"
                        else f"{v:.2f}"
                    )
                )
            )
        )
    )
    fmt_p = (
        lambda v: (
            "n/a"
            if (v is None or pd.isna(v))
            else f"{v * 100:.2f}%"
        )
    )
    fmt_ps = (
        lambda v: (
            "n/a"
            if (v is None or pd.isna(v))
            else f"{v * 100:+.2f}%"
        )
    )

    def mk_col(m, s, b_idx, c_idx):
        dash = "-"
        gc = (
            lambda idx: (
                m["Beta"][idx]
                if (
                    m is not None
                    and idx is not None
                    and len(m["Beta"]) > idx
                )
                else None
            )
        )
        return {
            "Fitted": fit_txt(m),
            "T1": (
                dash
                if m is None
                else fmt_n(
                    np.exp(m["Beta"][0])
                    * cfg["CostUnitScale"],
                    "#,##0.00",
                )
            ),
            "B": (
                dash
                if gc(b_idx) is None
                else fmt_n(gc(b_idx), "0.000000")
            ),
            "BS": (
                dash
                if gc(b_idx) is None
                else fmt_p(2 ** gc(b_idx))
            ),
            "C": (
                dash
                if gc(c_idx) is None
                else fmt_n(gc(c_idx), "0.000000")
            ),
            "CS": (
                dash
                if gc(c_idx) is None
                else fmt_p(2 ** gc(c_idx))
            ),
            "R2": (
                dash
                if m is None
                else fmt_n(gs(s, "R2"), "0.0000")
            ),
            "Adj": (
                dash
                if m is None
                else fmt_n(gs(s, "Adj"), "0.0000")
            ),
            "SEE": (
                dash
                if m is None
                else fmt_n(gs(s, "SEE"), "0.0000")
            ),
            "CV": dash if m is None else fmt_p(gs(s, "CV")),
            "MAPE": (
                dash
                if m is None
                else fmt_p(gs(s, "MAPE"))
            ),
            "Bias": (
                dash
                if m is None
                else fmt_ps(gs(s, "Bias"))
            ),
            "AICc": (
                dash
                if m is None
                else fmt_n(gs(s, "AICc"), "0.00")
            ),
            "DAI": (
                dash
                if m is None
                else fmt_n(d_aicc(s), "0.00")
            ),
            "T": (
                dash
                if m is None
                else fmt_n(gs(s, "T"), "0.00")
            ),
        }

    col_lc = mk_col(ctx["mdl_lc"], s_lc, 1, None)
    col_rt = mk_col(ctx["mdl_rt"], s_rt, None, 1)
    col_br = mk_col(ctx["mdl_lcr"], s_lcr, 1, 2)

    def r5(item, val, a, b, c):
        return {
            "Item": item,
            "Value": val,
            "LC": a,
            "Rate": b,
            "LC+Rate": c,
        }

    rows = [
        r5("Run ID", run_id, "", "", ""),
        r5("Program", program, "", "", ""),
        r5("Run label", run_label, "", "", ""),
        *[
            r5(item, value, "", "", "")
            for item, value in provenance(cfg).items()
        ],
        r5("Cost basis", cost_basis_txt, "", "", ""),
        r5("Source table", cfg["AnalogyTableName"], "", "", ""),
        r5("Analogy lots in fit", str(n_keep), "", "", ""),
        r5(
            "Quantity-only lots (units held, not fit)",
            str(n_unit - n_keep),
            "",
            "",
            "",
        ),
        r5("SD(ln qty)", fmt_n(rate_sd, "0.0000"), "", "", ""),
        r5(
            "Rate models",
            (
                "enabled"
                if rate_ok
                else f"gated off ({rate_why_not})"
            ),
            "",
            "",
            "",
        ),
        r5("", "", "", "", ""),
        r5(
            "Fitted",
            "",
            col_lc["Fitted"],
            col_rt["Fitted"],
            col_br["Fitted"],
        ),
        r5(
            "SELECTED",
            "",
            "YES" if sel == "LC" else "",
            "YES" if sel == "Rate" else "",
            "YES" if sel == "LC+Rate" else "",
        ),
        r5(
            "T1 ($K)",
            "",
            col_lc["T1"],
            col_rt["T1"],
            col_br["T1"],
        ),
        r5(
            "Learning exponent (b)",
            "",
            col_lc["B"],
            col_rt["B"],
            col_br["B"],
        ),
        r5(
            "Learning curve slope",
            "",
            col_lc["BS"],
            col_rt["BS"],
            col_br["BS"],
        ),
        r5(
            "Rate exponent (c)",
            "",
            col_lc["C"],
            col_rt["C"],
            col_br["C"],
        ),
        r5(
            "Rate slope",
            "",
            col_lc["CS"],
            col_rt["CS"],
            col_br["CS"],
        ),
        r5(
            "R2 (log)",
            "",
            col_lc["R2"],
            col_rt["R2"],
            col_br["R2"],
        ),
        r5(
            "Adj R2",
            "",
            col_lc["Adj"],
            col_rt["Adj"],
            col_br["Adj"],
        ),
        r5(
            "SEE (log)",
            "",
            col_lc["SEE"],
            col_rt["SEE"],
            col_br["SEE"],
        ),
        r5(
            "CV",
            "",
            col_lc["CV"],
            col_rt["CV"],
            col_br["CV"],
        ),
        r5(
            "MAPE",
            "",
            col_lc["MAPE"],
            col_rt["MAPE"],
            col_br["MAPE"],
        ),
        r5(
            "Mean bias",
            "",
            col_lc["Bias"],
            col_rt["Bias"],
            col_br["Bias"],
        ),
        r5(
            "AICc",
            "",
            col_lc["AICc"],
            col_rt["AICc"],
            col_br["AICc"],
        ),
        r5(
            "dAICc",
            "",
            col_lc["DAI"],
            col_rt["DAI"],
            col_br["DAI"],
        ),
        r5(
            "t (rate coefficient)",
            "",
            col_lc["T"],
            col_rt["T"],
            col_br["T"],
        ),
        r5(
            "Selection basis",
            "",
            sel_note if sel in ("LC", None) else "",
            sel_note if sel == "Rate" else "",
            sel_note if sel == "LC+Rate" else "",
        ),
    ]

    return pd.DataFrame(rows)


# ============================================================================
# 5. QUERY 3: FIT CHART DATA ENGINE
# ============================================================================
def generate_fit_chart_data(
    models_context: dict,
) -> pd.DataFrame:
    """Compute the fit chart evaluation data matching M code Fit Chart Data."""
    ctx = models_context
    cfg = ctx["cfg"]

    fit_q = ctx["fit_q"]
    fit_c = ctx["fit_c"]
    fit_se = ctx["fit_se"]

    t1_lc, b_lc = ctx["t1_lc"], ctx["b_lc"]
    t1_rt, b_rt = ctx["t1_rt"], ctx["b_rt"]
    t1_br, b_br, c_br = (
        ctx["t1_br"],
        ctx["b_br"],
        ctx["c_br"],
    )

    rnd = (
        lambda v, d: (
            round(v, d)
            if (v is not None and pd.notna(v))
            else np.nan
        )
    )

    chart_rows = []
    for k in range(len(fit_q)):
        lot_no = k + 1
        a_qty = int(fit_q[k])
        first_u = int(fit_se[k]["S"])
        last_u = int(fit_se[k]["E"])
        actual = rnd(fit_c[k] * cfg["CostUnitScale"], 2)

        # LC calculations
        lc_mid = rnd(
            lmp_func(first_u, last_u, a_qty, b_lc), 4
        )
        lc_est = (
            rnd(
                t1_lc
                * (
                    lmp_func(first_u, last_u, a_qty, b_lc)
                    ** b_lc
                )
                * cfg["CostUnitScale"],
                2,
            )
            if pd.notna(t1_lc)
            else np.nan
        )
        lc_res = (
            rnd(actual - lc_est, 2)
            if pd.notna(lc_est)
            else np.nan
        )
        lc_resp = (
            rnd(((actual / lc_est) - 1.0) * 100, 2)
            if (pd.notna(lc_est) and lc_est != 0)
            else np.nan
        )

        # Rate calculations (against Analogy Lot Quantity)
        rt_est = (
            rnd(
                t1_rt
                * (a_qty**b_rt)
                * cfg["CostUnitScale"],
                2,
            )
            if pd.notna(t1_rt)
            else np.nan
        )
        rt_res = (
            rnd(actual - rt_est, 2)
            if pd.notna(rt_est)
            else np.nan
        )
        rt_resp = (
            rnd(((actual / rt_est) - 1.0) * 100, 2)
            if (pd.notna(rt_est) and rt_est != 0)
            else np.nan
        )

        # LC+Rate calculations (True fitted value with rate term)
        lcr_mid = rnd(
            lmp_func(first_u, last_u, a_qty, b_br), 4
        )
        lcr_est = (
            rnd(
                t1_br
                * (
                    lmp_func(first_u, last_u, a_qty, b_br)
                    ** b_br
                )
                * (a_qty**c_br)
                * cfg["CostUnitScale"],
                2,
            )
            if pd.notna(t1_br)
            else np.nan
        )
        lcr_res = (
            rnd(actual - lcr_est, 2)
            if pd.notna(lcr_est)
            else np.nan
        )
        lcr_resp = (
            rnd(((actual / lcr_est) - 1.0) * 100, 2)
            if (pd.notna(lcr_est) and lcr_est != 0)
            else np.nan
        )

        chart_rows.append(
            {
                "Analogy Lot No.": lot_no,
                "Analogy Lot Quantity": a_qty,
                "First Unit in Lot": first_u,
                "Last Unit in Lot": last_u,
                "Actual AUC ($K)": actual,
                "LC Lot Midpoint": lc_mid,
                "LC Estimated AUC ($K)": lc_est,
                "LC Residual ($K)": lc_res,
                "LC Residual (%)": lc_resp,
                "Rate Estimated AUC ($K)": rt_est,
                "Rate Residual ($K)": rt_res,
                "Rate Residual (%)": rt_resp,
                "LC+Rate Lot Midpoint": lcr_mid,
                "LC+Rate Estimated AUC ($K)": lcr_est,
                "LC+Rate Residual ($K)": lcr_res,
                "LC+Rate Residual (%)": lcr_resp,
            }
        )

    chart_df = pd.DataFrame(chart_rows)
    return chart_df


# ============================================================================
# 6. EXCEL WORKBOOK & SCATTER CHART GENERATOR (FIXED OPENPYXL TICK MARKS)
# ============================================================================
#: A default Excel column is roughly 1.72 cm wide. Chart anchors are spaced off
#: this so widening a chart cannot silently overlap the one beside it.
_COL_CM = 1.72


def _nice_bounds(lo: float, hi: float, pad_frac: float = 0.15):
    """Axis bounds padded off the data and rounded to a readable step.

    Framing an axis on the data is what makes a tight band or a narrow
    distribution legible, but the raw min and max give labels like 298 and
    51.43M. Rounding outwards to a round step keeps both.
    """
    import math

    span = hi - lo
    if not math.isfinite(span) or span <= 0:
        span = abs(hi) or 1.0
    lo -= span * pad_frac
    hi += span * pad_frac
    width = hi - lo
    step = 10.0 ** math.floor(math.log10(width))
    for mult in (1, 2, 2.5, 5, 10):
        if width / (step * mult) <= 8:
            step *= mult
            break
    return math.floor(lo / step) * step, math.ceil(hi / step) * step


def _money_axis_fmt(lo: float, hi: float) -> str:
    """Excel number format for a money axis covering lo to hi.

    Each comma in an Excel format divides by a thousand. The decimal place
    appears only for a narrow spread, where rounding to whole units prints
    the same label several times in a row.
    """
    span = abs(hi - lo)
    if abs(hi) >= 1e9:
        return '#,##0.0,,,"B"' if span < 1e10 else '#,##0,,,"B"'
    if abs(hi) >= 1e6:
        return '#,##0.0,,"M"' if span < 1e7 else '#,##0,,"M"'
    if abs(hi) >= 1e3:
        return '#,##0.0,"K"' if span < 1e4 else '#,##0,"K"'
    return "#,##0"


def _money_short(value: float) -> str:
    """Compact money for a chart label: $250.0M rather than 250,000,000."""
    v = float(value)
    for cut, suffix in ((1e9, "B"), (1e6, "M"), (1e3, "K")):
        if abs(v) >= cut:
            return f"${v / cut:,.1f}{suffix}"
    return f"${v:,.0f}"


def _chart_anchor(index: int, width_cm: float, start_col: int = 2,
                  row: int = 10) -> str:
    """Anchor cell for the nth chart in a left-to-right row of charts."""
    step = int(width_cm / _COL_CM) + 4  # +4 columns of breathing room
    return f"{get_column_letter(start_col + index * step)}{row}"


def _format_chart(
    chart,
    title: str,
    x_title: str,
    y_title: str,
    width: float = 18,
    height: float = 11,
):
    """Apply the same axis, title and legend treatment to every chart.

    Two Excel quirks are handled here. openpyxl writes ``delete="1"`` onto a
    freshly created axis, which tells Excel to hide that axis completely, tick
    numbers and all; and a chart title defaults to overlaying the plot rather
    than sitting above it. The manual plot-area layout then reserves room on
    the left and bottom so the axis titles do not land on top of the tick
    numbers.
    """
    chart.title = title
    chart.style = 13
    chart.title.overlay = False

    chart.x_axis.title = x_title
    chart.y_axis.title = y_title

    for axis in (chart.x_axis, chart.y_axis):
        axis.delete = False
        axis.tickLblPos = "nextTo"
        axis.majorTickMark = "out"
        axis.minorTickMark = "none"
        axis.numFmt = "#,##0"

    chart.width = width
    chart.height = height

    # Keep the legend off the data.
    if chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False

    chart.layout = Layout(
        manualLayout=ManualLayout(
            xMode="edge",
            yMode="edge",
            x=0.11,
            y=0.13,
            w=0.86,
            h=0.68,
        )
    )


def save_complete_excel_workbook(
    filename: str,
    projections_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    chart_df: pd.DataFrame,
    risk_summary_df: pd.DataFrame | None = None,
    risk_intervals_df: pd.DataFrame | None = None,
    risk_scurve_df: pd.DataFrame | None = None,
):
    """Write the tables and embed native Excel scatter plots.

    The three risk frames are optional: they are present only when cost_core
    is installed and the risk analysis ran.
    """
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        summary_df.to_excel(
            writer, sheet_name="Analyst_Summary", index=False
        )
        projections_df.to_excel(
            writer,
            sheet_name="Estimate_Projections",
            index=False,
        )
        chart_df.to_excel(
            writer, sheet_name="Fit_Chart_Data", index=False
        )
        if risk_summary_df is not None:
            risk_summary_df.to_excel(
                writer, sheet_name="Risk_Summary", index=False
            )
        if risk_intervals_df is not None:
            risk_intervals_df.to_excel(
                writer, sheet_name="Risk_Intervals", index=False
            )
        if risk_scurve_df is not None:
            risk_scurve_df.to_excel(
                writer, sheet_name="Risk_SCurve", index=False
            )

    wb = openpyxl.load_workbook(filename)
    ws = wb["Fit_Chart_Data"]
    max_r = len(chart_df) + 1  # 1-indexed including header

    # Data labels inherit the source cell's number format, so format the
    # actual-AUC column and the labels come out with a thousands separator.
    for row in range(2, max_r + 1):
        ws.cell(row=row, column=5).number_format = "#,##0.00"

    fit_chart_width = 18

    def build_scatter_chart(
        title: str,
        x_col: int,
        actual_col: int,
        est_col: int,
        slot: int,
        x_axis_title_text: str,
        y_axis_title_text: str = "Unit Cost / AUC ($K)",
    ):
        chart = ScatterChart()
        _format_chart(
            chart,
            title,
            x_axis_title_text,
            y_axis_title_text,
            fit_chart_width,
            11,
        )

        x_values = Reference(
            ws, min_col=x_col, min_row=2, max_row=max_r
        )
        y_actual = Reference(
            ws, min_col=actual_col, min_row=1, max_row=max_r
        )
        y_est = Reference(
            ws, min_col=est_col, min_row=1, max_row=max_r
        )

        # Actuals: Markers with data labels showing numbers
        s_act = Series(
            values=y_actual,
            xvalues=x_values,
            title_from_data=True,
        )
        s_act.marker.symbol = "circle"
        s_act.marker.size = 7
        s_act.graphicalProperties.line.noFill = True

        # Print the actual AUC beside each marker.
        #
        # The attribute is dLbls. A Series has no `dataLabels` alias, unlike a
        # chart, so assigning to `series.dataLabels` sets a stray Python
        # attribute that never reaches the XML -- no error, no labels. The
        # other show flags are written explicitly because Excel treats an
        # absent flag as inherited rather than false.
        labels = DataLabelList()
        labels.showVal = True
        labels.showSerName = False
        labels.showCatName = False
        labels.showLegendKey = False
        labels.showBubbleSize = False
        labels.showPercent = False
        labels.dLblPos = "t"  # above the marker, clear of the fitted line
        # openpyxl can only write formatCode here, never sourceLinked, so
        # Excel falls back to the source cell's format. The AUC column is
        # formatted below to match.
        labels.numFmt = "#,##0.00"
        # 8pt, because the rate chart puts lots of equal quantity almost on
        # top of each other and full-size labels collide.
        labels.txPr = RichText(
            bodyPr=RichTextProperties(),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(sz=800)
                    ),
                    endParaRPr=CharacterProperties(sz=800),
                )
            ],
        )
        s_act.dLbls = labels

        # Estimates: Smooth Fitted Curve (No markers)
        s_est = Series(
            values=y_est,
            xvalues=x_values,
            title_from_data=True,
        )
        s_est.marker.symbol = "none"
        s_est.smooth = True

        chart.series.append(s_act)
        chart.series.append(s_est)
        ws.add_chart(chart, _chart_anchor(slot, fit_chart_width))

    # Chart 1: Learning Curve (LC) Fit -> X = LC Midpoint (Col 6), Y = Actual (Col 5) vs LC_Est (Col 7)
    build_scatter_chart(
        title="Learning Curve Fit: Actual vs Estimated AUC",
        x_col=6,
        actual_col=5,
        est_col=7,
        slot=0,
        x_axis_title_text="LC Lot Midpoint (Unit Number)",
        y_axis_title_text="Unit Cost / AUC ($K)",
    )

    # Chart 2: Rate Model Fit -> X = Lot Qty (Col 2), Y = Actual (Col 5) vs Rate_Est (Col 10)
    build_scatter_chart(
        title="Rate Model Fit: Actual vs Estimated AUC",
        x_col=2,
        actual_col=5,
        est_col=10,
        slot=1,
        x_axis_title_text="Analogy Lot Quantity (Units / Lot)",
        y_axis_title_text="Unit Cost / AUC ($K)",
    )

    # Chart 3: LC + Rate Fit -> X = LC+Rate Midpoint (Col 13), Y = Actual (Col 5) vs LCR_Est (Col 14)
    build_scatter_chart(
        title="LC+Rate Model Fit: Actual vs Estimated AUC",
        x_col=13,
        actual_col=5,
        est_col=14,
        slot=2,
        x_axis_title_text="LC+Rate Lot Midpoint (Unit Number)",
        y_axis_title_text="Unit Cost / AUC ($K)",
    )

    # Chart 4: the forecast with its prediction band, on the risk sheet.
    if risk_intervals_df is not None and len(risk_intervals_df):
        wsr = wb["Risk_Intervals"]
        last_r = len(risk_intervals_df) + 1

        # Find the columns by header rather than by position. cost_core owns
        # the shape of this frame, and it has changed once already.
        headers = {
            str(c.value): c.column
            for c in next(wsr.iter_rows(min_row=1, max_row=1))
        }
        x_name = next(
            (n for n in ("Last Unit in Lot", "Fiscal Year", "Lot Quantity")
             if n in headers),
            None,
        )
        y_names = [
            n for n in ("Unit Cost ($K)", "Unit Cost Lower", "Unit Cost Upper")
            if n in headers
        ]
        if x_name is None or len(y_names) < 2:
            wb.save(filename)
            return

        band = ScatterChart()
        _format_chart(
            band,
            "Forecast Unit Cost with Prediction Interval",
            x_name,
            "Unit Cost ($K)",
            20,
            11,
        )

        # Excel starts a value axis at zero, which squeezes a tight prediction
        # band into what looks like one thick line. Frame the axis on the band
        # itself so its width is actually readable.
        lo_y, hi_y = _nice_bounds(
            float(risk_intervals_df["Unit Cost Lower"].min()),
            float(risk_intervals_df["Unit Cost Upper"].max()),
        )
        band.y_axis.scaling.min = max(0.0, lo_y)
        band.y_axis.scaling.max = hi_y

        xs = Reference(
            wsr, min_col=headers[x_name], min_row=2, max_row=last_r
        )
        for name in y_names:
            dashed = name != "Unit Cost ($K)"
            s = Series(
                values=Reference(
                    wsr, min_col=headers[name], min_row=1, max_row=last_r
                ),
                xvalues=xs,
                title_from_data=True,
            )
            s.marker.symbol = "none" if dashed else "circle"
            s.smooth = False
            if dashed:
                s.graphicalProperties.line.dashStyle = "dash"
            band.series.append(s)
        wsr.add_chart(band, "P2")

    # Chart 5: the S-curve. Cost on the x axis, cumulative probability on the
    # y axis, which is the orientation everyone reads a P80 off.
    if risk_scurve_df is not None and len(risk_scurve_df):
        wss = wb["Risk_SCurve"]
        last_s = len(risk_scurve_df) + 1
        for row in range(2, last_s + 1):
            wss.cell(row=row, column=1).number_format = "0%"
            wss.cell(row=row, column=2).number_format = "#,##0"

        curve = ScatterChart()
        _format_chart(
            curve,
            "Cost S-Curve: Probability the Buy Comes In At or Below",
            "Buy Total ($)",
            "Cumulative Probability",
            20,
            11,
        )
        curve.y_axis.numFmt = "0%"
        # Whole dollars would give an axis of unreadable 9-digit labels that
        # collide; each comma is a division by a thousand in an Excel format.
        # The decimal place appears only for a narrow spread, where rounding
        # to whole units would print the same label several times over.
        lo_x = float(risk_scurve_df["Buy Total ($)"].min())
        hi_x = float(risk_scurve_df["Buy Total ($)"].max())
        curve.x_axis.numFmt = _money_axis_fmt(lo_x, hi_x)
        # The P50 and P80 markers are named in the legend rather than by data
        # labels beside them. Excel can only place a label immediately next to
        # its point, and the curve runs through the point, so on a steep
        # S-curve every available position puts the line through the text.
        # The legend sits below the plot where nothing can overlap it.
        if curve.legend is not None:
            curve.legend.position = "b"
            curve.legend.overlay = False
        # Probability is bounded, and Excel otherwise draws an axis to 120%.
        curve.y_axis.scaling.min = 0
        curve.y_axis.scaling.max = 1
        # Zoom to where the distribution actually sits, rather than showing
        # 500M of empty space because Excel starts every axis at zero.
        nice_lo, nice_hi = _nice_bounds(lo_x, hi_x)
        curve.x_axis.scaling.min = max(0.0, nice_lo)
        curve.x_axis.scaling.max = nice_hi
        s = Series(
            values=Reference(wss, min_col=1, min_row=1, max_row=last_s),
            xvalues=Reference(wss, min_col=2, min_row=2, max_row=last_s),
            title_from_data=True,
        )
        s.marker.symbol = "none"
        s.smooth = True
        s.tx = SeriesLabel(v="Buy total distribution")
        curve.series.append(s)

        # Call out P50 and P80. Each is its own one-point series, named for
        # the percentile and its cost, so the legend carries both while the
        # marker shows where it falls on the curve.
        pct = risk_scurve_df["Percentile"].round(4)
        # Left to its own palette Excel gives these two more shades of the
        # curve's green, so they disappear into it. Fixed colours from the
        # Okabe-Ito set, which stays distinguishable for the common colour
        # vision deficiencies, plus a different shape each so the two are
        # still telling apart in greyscale or on a photocopy.
        marks = [
            ("P50", 0.50, "0072B2", "circle"),      # blue
            ("P80", 0.80, "D55E00", "diamond"),     # vermillion
        ]
        for i, (name, level, colour, symbol) in enumerate(marks):
            hit = risk_scurve_df.loc[pct == round(level, 4), "Buy Total ($)"]
            if hit.empty:
                continue
            cost = float(hit.iloc[0])
            y_col = 4 + i * 2  # D, then F
            x_col = y_col + 1  # E, then G
            wss.cell(row=1, column=y_col, value=f"{name}  {_money_short(cost)}")
            wss.cell(row=2, column=y_col, value=level).number_format = "0%"
            wss.cell(row=1, column=x_col, value=f"{name} cost")
            wss.cell(
                row=2, column=x_col, value=cost
            ).number_format = "#,##0"

            mark = Series(
                values=Reference(wss, min_col=y_col, min_row=1, max_row=2),
                xvalues=Reference(wss, min_col=x_col, min_row=2, max_row=2),
                title_from_data=True,
            )
            mark.marker.symbol = symbol
            mark.marker.size = 11
            # Filled in the marker's own colour with a white keyline, so it
            # reads as a marker sitting on the curve rather than a kink in it.
            marker_style = GraphicalProperties(solidFill=colour)
            marker_style.line.solidFill = "FFFFFF"
            marker_style.line.width = 19050  # 1.5pt, in EMU
            mark.marker.graphicalProperties = marker_style
            # No connecting line: these are single points.
            mark.graphicalProperties.line.noFill = True
            curve.series.append(mark)

        wss.add_chart(curve, "I2")

    wb.save(filename)



# ============================================================================
# 7. GUI DATA ENTRY (tkinter - ships with Python, no install needed)
# ============================================================================
import os
import sys
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# Optional. Without it every deterministic feature still works; only the
# Risk tab goes dark.
try:
    import risk
except Exception:  # pragma: no cover - depends on the install
    risk = None

# The WBS roll-up. Its own risk half needs cost_core, but the deterministic
# roll-up does not, so this stays useful either way.
try:
    import wbs
except Exception:  # pragma: no cover - depends on the install
    wbs = None

# Demo data for the "Load Example" buttons. These numbers are invented,
# generated from an 88% learning curve with a 93% rate slope on a $1,200K
# first unit, plus a little scatter.
#
# Every lot is a different size, spanning 5 to 50 units. That matters: with
# repeated quantities there is nothing for a rate term to regress against, and
# the Rate chart collapses into a cluster of points at two or three values.
# Here all three models fit, the rate coefficient clears the significance
# gate, and LC+Rate is selected on its merits.
EXAMPLE_ANALOGY = [
    ("2015", "5", "857.91"),
    ("2016", "9", "645.57"),
    ("2017", "14", "531.74"),
    ("2018", "22", "437.51"),
    ("2019", "34", "380.10"),
    ("2020", "50", "332.21"),
]

# The buy grows and then tapers, so the rate term does visible work: the last
# lot is small, and its unit cost comes back up rather than continuing down.
EXAMPLE_ESTIMATE = [
    ("2028", "12", "1.15"),
    ("2029", "20", "1.15"),
    ("2030", "30", "1.15"),
    ("2031", "40", "1.15"),
    ("2032", "25", "1.15"),
    ("2033", "10", "1.15"),
]


def parse_float(text: str) -> float:
    """Parse a number, tolerating $ signs and thousands separators."""
    return float(text.replace("$", "").replace(",", "").strip())


def split_row(line: str) -> list[str]:
    """Split a pasted row on tabs, commas, or runs of spaces."""
    line = line.rstrip("\r")
    if "\t" in line:
        parts = line.split("\t")
    elif "," in line:
        parts = line.split(",")
    else:
        parts = line.split()
    return [p.strip() for p in parts]


def default_output_dir() -> str:
    """Somewhere guaranteed writable: the script's folder, else Documents."""
    try:
        here = os.path.dirname(os.path.abspath(__file__))
        probe = os.path.join(here, ".__writetest")
        with open(probe, "w") as fh:
            fh.write("x")
        os.remove(probe)
        return here
    except Exception:
        return os.path.join(os.path.expanduser("~"), "Documents")


class LotGrid(ttk.Frame):
    """A small scrollable spreadsheet of Entry widgets."""

    def __init__(self, parent, headers: list[str], widths: list[int]):
        super().__init__(parent)
        self.headers = headers
        self.widths = widths
        self.rows: list[list[tk.Entry]] = []

        head = ttk.Frame(self)
        head.pack(fill="x", padx=(4, 0))
        ttk.Label(head, text="#", width=4, anchor="center").grid(
            row=0, column=0, padx=1
        )
        self.head_labels: list[ttk.Label] = []
        self.hidden: set[int] = set()
        for i, (h, w) in enumerate(zip(headers, widths)):
            lbl = ttk.Label(
                head, text=h, width=w, anchor="center", style="Head.TLabel"
            )
            lbl.grid(row=0, column=i + 1, padx=1)
            self.head_labels.append(lbl)

        canvas_wrap = ttk.Frame(self)
        canvas_wrap.pack(fill="both", expand=True)
        self.canvas = tk.Canvas(canvas_wrap, highlightthickness=0, height=240)
        scroll = ttk.Scrollbar(
            canvas_wrap, orient="vertical", command=self.canvas.yview
        )
        self.canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.body = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.body, anchor="nw")
        self.body.bind(
            "<Configure>",
            lambda e: self.canvas.configure(
                scrollregion=self.canvas.bbox("all")
            ),
        )
        self.canvas.bind_all("<MouseWheel>", self._on_wheel)

    def set_headers(self, headers: list[str]):
        """Relabel the columns. An empty heading hides that column.

        The grid is one widget reused by every element, so what a column
        means changes with the element kind. A column that does not apply is
        hidden rather than left showing a heading the tool will not read.
        """
        self.headers = list(headers)
        self.hidden = {i for i, h in enumerate(headers) if not h}
        for i, (lbl, h) in enumerate(zip(self.head_labels, headers)):
            lbl.config(text=h)
            if i in self.hidden:
                lbl.grid_remove()
            else:
                lbl.grid()
        for entries in self.rows:
            self._apply_hidden(entries)

    def _apply_hidden(self, entries):
        for i, e in enumerate(entries):
            if i in self.hidden:
                e.delete(0, tk.END)
                e.grid_remove()
            else:
                e.grid()

    def _on_wheel(self, event):
        try:
            self.canvas.yview_scroll(int(-event.delta / 120), "units")
        except Exception:
            pass

    def add_row(self, values: tuple = None):
        r = len(self.rows)
        ttk.Label(self.body, text=str(r + 1), width=4, anchor="center").grid(
            row=r, column=0, padx=1, pady=1
        )
        entries = []
        for i, w in enumerate(self.widths):
            e = tk.Entry(self.body, width=w, justify="right")
            e.grid(row=r, column=i + 1, padx=1, pady=1)
            if values and i < len(values):
                e.insert(0, values[i])
            e.bind("<Control-v>", self._on_paste)
            e.bind("<Return>", lambda ev: self.add_row())
            entries.append(e)
        self._apply_hidden(entries)
        self.rows.append(entries)
        self.body.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        return entries

    def _on_paste(self, event):
        """Multi-line clipboard content fills the grid from this row down."""
        try:
            data = self.clipboard_get()
        except tk.TclError:
            return None
        if "\n" not in data.strip():
            return None  # single value: let Tk paste normally

        widget = event.widget
        start = 0
        for idx, row in enumerate(self.rows):
            if widget in row:
                start = idx
                break

        lines = [ln for ln in data.split("\n") if ln.strip()]
        for offset, line in enumerate(lines):
            target = start + offset
            while target >= len(self.rows):
                self.add_row()
            parts = split_row(line)
            for col, entry in enumerate(self.rows[target]):
                entry.delete(0, tk.END)
                if col < len(parts):
                    entry.insert(0, parts[col])
        return "break"

    def delete_last(self):
        if not self.rows:
            return
        for w in self.body.grid_slaves(row=len(self.rows) - 1):
            w.destroy()
        self.rows.pop()
        self.body.update_idletasks()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def clear(self):
        while self.rows:
            self.delete_last()

    def load(self, data: list):
        self.clear()
        for values in data:
            self.add_row(values)

    def get_rows(self) -> list[list[str]]:
        """Non-empty rows as raw strings."""
        out = []
        for row in self.rows:
            vals = [e.get().strip() for e in row]
            if any(v for v in vals):
                out.append(vals)
        return out


#: Marker and schema version for a saved run. The version is bumped only when
#: an older file can no longer be read as written.
RUN_FORMAT = "lot-cost-model-run"
#: 2 added the WBS elements list. A version 1 file still loads, as a single
#: element. Bumping means an older build refuses a version 2 file outright
#: rather than opening it and silently keeping only one of its elements.
RUN_FORMAT_VERSION = 2
RUN_SUFFIX = ".lotrun.json"


class RunFileError(Exception):
    """A saved run could not be read."""


def read_run_file(path: str) -> dict:
    """Load and validate a saved run.

    Raises:
        RunFileError: If the file is not a saved run, or was written by a
            newer format than this build understands.
    """
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
    except json.JSONDecodeError as exc:
        raise RunFileError(
            f"{os.path.basename(path)} is not a saved run: {exc}"
        ) from exc
    except OSError as exc:
        raise RunFileError(f"Could not open {path}: {exc}") from exc

    if not isinstance(data, dict) or data.get("format") != RUN_FORMAT:
        raise RunFileError(
            f"{os.path.basename(path)} is not a lot cost model run file."
        )
    version = data.get("format_version", 0)
    if not isinstance(version, int) or version > RUN_FORMAT_VERSION:
        raise RunFileError(
            f"{os.path.basename(path)} was written in run format {version}, "
            f"and this build reads up to {RUN_FORMAT_VERSION}. Update the "
            "tool to open it."
        )
    return data


def _scroll(parent, tree):
    """Pack a treeview with a vertical scrollbar beside it."""
    bar = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=bar.set)
    bar.pack(side="right", fill="y")
    tree.pack(fill="both", expand=True)
    return bar


class LotCostApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Lot Cost Model - Learning Curve / Rate Analysis")
        self.geometry("980x760")
        self.minsize(860, 640)

        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Head.TLabel", font=("Segoe UI", 9, "bold"))
        style.configure("Title.TLabel", font=("Segoe UI", 13, "bold"))
        style.configure("Sub.TLabel", foreground="#555")

        ttk.Label(
            self, text="Lot Cost Model", style="Title.TLabel"
        ).pack(anchor="w", padx=12, pady=(10, 0))
        ttk.Label(
            self,
            text=(
                "Enter historical analogy lots and forecast lots, then run the "
                "model. Paste directly from Excel with Ctrl+V."
            ),
            style="Sub.TLabel",
        ).pack(anchor="w", padx=12, pady=(0, 8))

        # One entry per WBS element: its own analogy history and its own
        # share of the buy. Tabs 1 to 5 always show the selected element;
        # tab 6 is the whole programme.
        self.elements: list[dict] = []
        self.current_element = 0
        self._switching = False

        self._build_element_bar()
        self._build_kind_bar(self)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12)
        self.tab_analogy = ttk.Frame(nb)
        self.tab_estimate = ttk.Frame(nb)
        self.tab_run = ttk.Frame(nb)
        self.tab_results = ttk.Frame(nb)
        self.tab_risk = ttk.Frame(nb)
        self.tab_program = ttk.Frame(nb)
        nb.add(self.tab_analogy, text="  1. Analogy Lots  ")
        nb.add(self.tab_estimate, text="  2. Estimate Lots  ")
        nb.add(self.tab_run, text="  3. Run Info & Settings  ")
        nb.add(self.tab_results, text="  4. Results  ")
        nb.add(self.tab_risk, text="  5. Risk & Intervals  ")
        nb.add(self.tab_program, text="  6. Program Roll-up  ")
        self.nb = nb

        self.risk_result = None
        self.program_result = None

        self.run_path: str | None = None

        self._build_analogy()
        self._build_estimate()
        self._build_runinfo()
        self._build_results()
        self._build_risk()
        self._build_program()
        self._build_actionbar()
        self._build_menu()

        # Start with a single unnamed element, so a one-element estimate
        # behaves exactly as it did before any of this existed.
        self.elements = [self._blank_element("Element 1")]
        self._refresh_element_list()

    # -- tabs ---------------------------------------------------------------
    def _build_analogy(self):
        f = self.tab_analogy
        ttk.Label(
            f,
            text=(
                "Historical lots used to fit the curve. Need at least 3 lots "
                "with a unit cost.\n"
                "Leave AUC blank for a quantity-only lot (its units count "
                "toward learning, but it is not fit)."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 6))

        self.grid_analogy = LotGrid(
            f,
            ["Fiscal Year", "Lot Quantity", "Unit Cost AUC ($K)"],
            [14, 14, 20],
        )
        self.grid_analogy.pack(fill="both", expand=True, padx=8)

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Button(
            bar, text="Add Row", command=self.grid_analogy.add_row
        ).pack(side="left")
        ttk.Button(
            bar, text="Delete Last Row", command=self.grid_analogy.delete_last
        ).pack(side="left", padx=4)
        ttk.Button(
            bar, text="Clear", command=self.grid_analogy.clear
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Load Example",
            command=lambda: self.grid_analogy.load(EXAMPLE_ANALOGY),
        ).pack(side="right")

        for _ in range(6):
            self.grid_analogy.add_row()

    def _build_estimate(self):
        f = self.tab_estimate
        ttk.Label(
            f,
            text=(
                "Forecast lots to be costed. Complexity Factor is optional; a "
                "blank one carries the\nprevious lot's value forward (1.0 to "
                "start)."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 6))

        self.grid_estimate = LotGrid(
            f,
            ["Fiscal Year", "Lot Quantity", "Complexity Factor"],
            [14, 14, 20],
        )
        self.grid_estimate.pack(fill="both", expand=True, padx=8)

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Button(
            bar, text="Add Row", command=self.grid_estimate.add_row
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Delete Last Row",
            command=self.grid_estimate.delete_last,
        ).pack(side="left", padx=4)
        ttk.Button(
            bar, text="Clear", command=self.grid_estimate.clear
        ).pack(side="left")
        ttk.Button(
            bar,
            text="Load Example",
            command=lambda: self.grid_estimate.load(EXAMPLE_ESTIMATE),
        ).pack(side="right")

        for _ in range(8):
            self.grid_estimate.add_row()

    def _build_runinfo(self):
        f = self.tab_run
        box = ttk.LabelFrame(f, text="Run Info")
        box.pack(fill="x", padx=8, pady=10)

        self.var_runid = tk.StringVar(value=SETTINGS["DefaultRunID"])
        self.var_program = tk.StringVar(value=SETTINGS["DefaultProgram"])
        self.var_label = tk.StringVar(value=SETTINGS["DefaultRunLabel"])
        self.var_baseyear = tk.StringVar(value="")

        fields = [
            ("Run ID", self.var_runid, ""),
            ("Program", self.var_program, ""),
            ("Run label", self.var_label, ""),
            ("Base year ($)", self.var_baseyear, "blank = not stated"),
        ]
        for r, (lbl, var, hint) in enumerate(fields):
            ttk.Label(box, text=lbl + ":").grid(
                row=r, column=0, sticky="e", padx=8, pady=5
            )
            ttk.Entry(box, textvariable=var, width=42).grid(
                row=r, column=1, sticky="w", pady=5
            )
            if hint:
                ttk.Label(box, text=hint, style="Sub.TLabel").grid(
                    row=r, column=2, sticky="w", padx=8
                )

        box2 = ttk.LabelFrame(f, text="Model Settings")
        box2.pack(fill="x", padx=8, pady=6)

        self.var_costscale = tk.StringVar(
            value=str(SETTINGS["CostUnitScale"])
        )
        self.var_totalscale = tk.StringVar(value=str(SETTINGS["TotalScale"]))
        self.var_defaultcf = tk.StringVar(value=str(SETTINGS["DefaultCF"]))
        self.var_tgate = tk.StringVar(value=str(SETTINGS["TGate"]))
        self.var_fitprior = tk.StringVar(
            value=str(SETTINGS["FitPriorUnits"])
        )
        self.var_fcstprior = tk.StringVar(
            value=str(SETTINGS["FcstPriorUnits"])
        )
        self.var_legacy_rate = tk.BooleanVar(
            value=SETTINGS["LegacyRateOmission"]
        )

        s_fields = [
            ("Cost unit scale", self.var_costscale, "1 = $K, 1000 = dollars"),
            ("Total scale", self.var_totalscale, "applied on top, for totals"),
            ("Default complexity", self.var_defaultcf, "used if none given"),
            ("t-gate", self.var_tgate, "significance cutoff for rate term"),
            ("Prior units (fit)", self.var_fitprior, "units built before lot 1"),
            ("Prior units (forecast)", self.var_fcstprior, ""),
        ]
        for r, (lbl, var, hint) in enumerate(s_fields):
            ttk.Label(box2, text=lbl + ":").grid(
                row=r, column=0, sticky="e", padx=8, pady=4
            )
            ttk.Entry(box2, textvariable=var, width=16).grid(
                row=r, column=1, sticky="w", pady=4
            )
            if hint:
                ttk.Label(box2, text=hint, style="Sub.TLabel").grid(
                    row=r, column=2, sticky="w", padx=8
                )
        ttk.Checkbutton(
            box2,
            text=(
                "Legacy rate omission: drop the rate term when projecting "
                "(overstates Rate and LC+Rate; for reconciling old workbooks "
                "only)"
            ),
            variable=self.var_legacy_rate,
        ).grid(
            row=len(s_fields), column=0, columnspan=3, sticky="w", padx=8,
            pady=6,
        )

    def _build_results(self):
        f = self.tab_results
        self.lbl_result = ttk.Label(
            f,
            text=(
                "No run yet. These tabs show one element at a time: fill in "
                "its lots and click Run Model, or run the program roll-up on "
                "tab 6, which fills this in for the element you have selected."
            ),
            style="Sub.TLabel",
        )
        self.lbl_result.pack(anchor="w", padx=8, pady=8)

        cols = ("Item", "Value", "LC", "Rate", "LC+Rate")
        self.tree = ttk.Treeview(f, columns=cols, show="headings", height=22)
        widths = (250, 330, 120, 120, 120)
        for c, w in zip(cols, widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="w")
        vs = ttk.Scrollbar(f, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vs.set)
        vs.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.tree.tag_configure("sel", background="#dff0d8")

    # -- program roll-up ----------------------------------------------------
    def _build_program(self):
        f = self.tab_program
        if wbs is None:
            ttk.Label(
                f,
                text=(
                    "Program roll-up unavailable: wbs.py is not next to this "
                    "script."
                ),
                style="Sub.TLabel",
            ).pack(anchor="w", padx=12, pady=14)
            self.tree_prog_elements = None
            self.tree_prog_summary = None
            return

        ttk.Label(
            f,
            text=(
                "Every element priced on its own curve, then added up. The "
                "cost before risk stands on its own:\n"
                "risk is optional, reported separately, and never folded into "
                "the estimate above it."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 4))

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=4)
        ttk.Button(
            bar, text="Run Program Roll-up", command=self.run_program
        ).pack(side="left")
        self.var_prog_risk = tk.BooleanVar(value=bool(wbs.RISK_AVAILABLE))
        ttk.Checkbutton(
            bar,
            text="Also apply risk (correlated Monte Carlo)",
            variable=self.var_prog_risk,
            state="normal" if wbs.RISK_AVAILABLE else "disabled",
        ).pack(side="left", padx=10)
        ttk.Label(bar, text="Correlation:").pack(side="left", padx=(10, 2))
        self.var_prog_rho = tk.StringVar(
            value=f"{wbs.DEFAULT_CORRELATION:.2f}"
        )
        ttk.Entry(bar, textvariable=self.var_prog_rho, width=6).pack(
            side="left"
        )
        self.var_prog_status = tk.StringVar(value="Not run yet.")
        ttk.Label(
            bar, textvariable=self.var_prog_status, style="Sub.TLabel"
        ).pack(side="left", padx=12)

        self._build_program_views(f)

    def _build_program_views(self, parent):
        """Roll-up, tornado, influence and sensitivity, as inner tabs."""
        inner = ttk.Notebook(parent)
        inner.pack(fill="both", expand=True)
        self.prog_views = inner

        # -- roll-up ---------------------------------------------------------
        page = ttk.Frame(inner)
        inner.add(page, text="  Roll-up  ")
        panes = ttk.Panedwindow(page, orient="vertical")
        panes.pack(fill="both", expand=True)

        top = ttk.Frame(panes)
        cols = (
            "WBS Element", "Model", "Lots", "T1 ($K)", "Units",
            "Cost Before Risk ($)", "Share", "P80 With Risk ($)",
        )
        self.tree_prog_elements = ttk.Treeview(
            top, columns=cols, show="headings", height=7
        )
        for c, w in zip(cols, (180, 75, 45, 85, 65, 155, 65, 145)):
            self.tree_prog_elements.heading(c, text=c)
            self.tree_prog_elements.column(
                c, width=w, anchor="w" if c == "WBS Element" else "e"
            )
        _scroll(top, self.tree_prog_elements)
        self.tree_prog_elements.tag_configure(
            "total", background="#dff0d8", font=("Segoe UI", 9, "bold")
        )
        panes.add(top, weight=2)

        bot = ttk.Frame(panes)
        self.tree_prog_summary = ttk.Treeview(
            bot, columns=("Item", "Value"), show="headings", height=10
        )
        self.tree_prog_summary.heading("Item", text="Item")
        self.tree_prog_summary.heading("Value", text="Value")
        self.tree_prog_summary.column("Item", width=280, anchor="w")
        self.tree_prog_summary.column("Value", width=640, anchor="w")
        _scroll(bot, self.tree_prog_summary)
        self.tree_prog_summary.tag_configure(
            "head", font=("Segoe UI", 9, "bold")
        )
        panes.add(bot, weight=3)

        # -- funding by fiscal year ------------------------------------------
        page = ttk.Frame(inner)
        inner.add(page, text="  Funding by FY  ")
        ttk.Label(
            page,
            text=(
                "The budget view: one row per fiscal year whatever the lot "
                "structure, so two lots awarded in one year are one year of "
                "funding. Every lot's cost sits in the year it is awarded."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=6, pady=(6, 4))
        self.tree_prog_fy = ttk.Treeview(
            page, columns=("a",), show="headings", height=12
        )
        _scroll(page, self.tree_prog_fy)

        # -- tornado ---------------------------------------------------------
        page = ttk.Frame(inner)
        inner.add(page, text="  Tornado  ")
        ttk.Label(
            page,
            text=(
                "Share of program variance, largest first. This is not the "
                "same ranking as size: an element that is only moderately\n"
                "variable but moves with everything else contributes more "
                "than its cost share suggests. Needs risk applied."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=6, pady=(6, 4))
        cols = ("Component", "Share of Variance", "Std Dev ($)")
        self.tree_prog_tornado = ttk.Treeview(
            page, columns=cols, show="headings", height=10
        )
        for c, w in zip(cols, (300, 160, 160)):
            self.tree_prog_tornado.heading(c, text=c)
            self.tree_prog_tornado.column(
                c, width=w, anchor="w" if c == "Component" else "e"
            )
        _scroll(page, self.tree_prog_tornado)

        # -- influence -------------------------------------------------------
        page = ttk.Frame(inner)
        inner.add(page, text="  Influence  ")
        ttk.Label(
            page,
            text=(
                "Which analogy lot is carrying each element's fit. At six "
                "lots one can set the slope while every summary\nstatistic "
                "still looks healthy. Flags, not verdicts: the largest or "
                "smallest lot has high leverage by construction."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=6, pady=(6, 4))
        cols = ("WBS Element", "Lot", "Qty", "% Error", "Leverage", "Cook's D")
        self.tree_prog_influence = ttk.Treeview(
            page, columns=cols, show="headings", height=12
        )
        for c, w in zip(cols, (200, 120, 70, 90, 90, 90)):
            self.tree_prog_influence.heading(c, text=c)
            self.tree_prog_influence.column(
                c, width=w, anchor="w" if c == "WBS Element" else "e"
            )
        _scroll(page, self.tree_prog_influence)
        self.tree_prog_influence.tag_configure("flag", background="#fde9d9")

        # -- buy sensitivity -------------------------------------------------
        page = ttk.Frame(inner)
        inner.add(page, text="  Buy Sensitivity  ")
        ttk.Label(
            page,
            text=(
                "Reprice the whole program at other buy sizes. Every "
                "element's quantities scale together, so engines per\n"
                "aircraft and spares keep their proportion. Buying fewer "
                "pushes unit cost up, which is the rate term at work."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=6, pady=(6, 4))
        row = ttk.Frame(page)
        row.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Label(row, text="Buy multipliers:").pack(side="left")
        self.var_prog_factors = tk.StringVar(value="0.6, 0.8, 1.0, 1.2, 1.5")
        ttk.Entry(row, textvariable=self.var_prog_factors, width=30).pack(
            side="left", padx=6
        )
        ttk.Button(
            row, text="Run Buy Sensitivity", command=self.run_sensitivity
        ).pack(side="left")
        self.tree_prog_sens = ttk.Treeview(
            page, columns=("a",), show="headings", height=10
        )
        _scroll(page, self.tree_prog_sens)
        self.tree_prog_sens.tag_configure("base", background="#dff0d8")

    def _show_program_extras(self, result):
        """Fill the funding, tornado and influence views from a roll-up."""
        annual = wbs.by_fiscal_year(result)
        cols = list(annual.columns)
        self.tree_prog_fy["columns"] = cols
        for c in cols:
            self.tree_prog_fy.heading(c, text=c)
            self.tree_prog_fy.column(
                c, width=70 if c in ("Fiscal Year", "Lots") else 150,
                anchor="e",
            )
        self.tree_prog_fy.delete(*self.tree_prog_fy.get_children())
        for _, r in annual.iterrows():
            values = []
            for c in cols:
                v = r[c]
                if c in ("Fiscal Year", "Lots"):
                    values.append(f"{int(v)}")
                elif "Share" in c:
                    values.append(f"{float(v):.1%}")
                else:
                    values.append(f"{float(v):,.2f}")
            self.tree_prog_fy.insert("", "end", values=values)

        self.tree_prog_tornado.delete(*self.tree_prog_tornado.get_children())
        if result.tornado is not None and len(result.tornado):
            for _, r in result.tornado.iterrows():
                self.tree_prog_tornado.insert(
                    "",
                    "end",
                    values=(
                        r.get("component", ""),
                        f"{float(r.get('variance_share', 0)):.1%}",
                        f"{float(r.get('std_dev', 0)):,.0f}",
                    ),
                )
        else:
            self.tree_prog_tornado.insert(
                "", "end",
                values=("Tick 'Also apply risk' to rank the elements.",
                        "", ""),
            )

        self.tree_prog_influence.delete(
            *self.tree_prog_influence.get_children()
        )
        table = wbs.influence_table(result)
        if table is None or table.empty:
            self.tree_prog_influence.insert(
                "", "end",
                values=("Influence diagnostics need cost_core installed.",
                        "", "", "", "", ""),
            )
            return
        cook_col = "Cook's D"
        for _, r in table.iterrows():
            flagged = bool(r.get("Influential", False)) or bool(
                r.get("High leverage", False)
            )
            self.tree_prog_influence.insert(
                "",
                "end",
                values=(
                    r.get("WBS Element", ""),
                    r.get("Lot", ""),
                    f"{float(r.get('Qty', 0)):,.0f}",
                    f"{float(r.get('% error', 0)):+.2f}%",
                    f"{float(r.get('Leverage', 0)):.3f}",
                    f"{float(r.get(cook_col, 0)):.3f}",
                ),
                tags=("flag",) if flagged else (),
            )

    def run_sensitivity(self):
        """Reprice the whole program at several buy sizes."""
        if wbs is None:
            return
        self.var_prog_status.set("Repricing at each buy size...")
        self.update_idletasks()
        try:
            factors = [
                float(x)
                for x in self.var_prog_factors.get().replace(";", ",").split(",")
                if x.strip()
            ]
            if not factors:
                raise ValueError("Give at least one buy multiplier.")
            frame = wbs.buy_profile_sensitivity(
                self.build_program(), factors, self._collect_overrides()
            )
        except (ValueError, wbs.ProgramError) as exc:
            messagebox.showerror("Buy sensitivity", str(exc))
            self.var_prog_status.set("Sensitivity did not run.")
            return
        except Exception as exc:
            messagebox.showerror(
                "Buy sensitivity failed", f"{type(exc).__name__}: {exc}"
            )
            self.var_prog_status.set("Sensitivity failed.")
            return

        self.sensitivity_result = frame
        cols = list(frame.columns)
        self.tree_prog_sens["columns"] = cols
        for c in cols:
            self.tree_prog_sens.heading(c, text=c)
            self.tree_prog_sens.column(
                c, width=160 if "$" in c else 120, anchor="e"
            )
        self.tree_prog_sens.delete(*self.tree_prog_sens.get_children())
        for _, r in frame.iterrows():
            values = []
            for c in cols:
                v = r[c]
                if c == "Buy Multiplier":
                    values.append(f"{float(v):.2f}x")
                elif "vs Baseline" in c:
                    values.append(f"{float(v):+.1%}")
                elif "$" in c:
                    values.append(f"{float(v):,.2f}")
                else:
                    values.append(f"{float(v):,.0f}")
            self.tree_prog_sens.insert(
                "", "end", values=values,
                tags=("base",) if float(r["Buy Multiplier"]) == 1.0 else (),
            )
        self.prog_views.select(3)
        self.var_prog_status.set(
            f"Repriced at {len(frame)} buy sizes; the baseline row is "
            "highlighted."
        )

    def build_program(self):
        """Assemble the programme from what is entered, without pricing it.

        The shared schedule comes from the first element that has one. A
        factor element has no lots of its own, so it does not contribute a
        schedule and does not have to match one.
        """
        self._capture_element()
        elements = []
        schedule: list[int] = []
        schedule_from = ""

        for el in self.elements:
            kind = el.get("kind", "fitted")
            estimate_rows = [r for r in el["estimate"] if any(r)]

            if kind == "factor":
                elements.append(
                    wbs.factor_of(
                        el["name"],
                        float(el.get("factor", 0.0)),
                        list(el.get("basis") or []) or None,
                    )
                )
                continue

            if not estimate_rows:
                raise ValueError(
                    f"{el['name']} has no forecast lots."
                    if kind == "fitted"
                    else f"{el['name']} has no costs yet. Click Phase a "
                    "total to spread one number across the years, or type "
                    "each year's cost in the Amount column on tab 2."
                )
            years, values, cf = self._estimate_columns(
                estimate_rows, el["name"]
            )
            if not schedule:
                schedule, schedule_from = years, el["name"]
            elif len(years) != len(schedule):
                raise ValueError(
                    f"{el['name']} has {len(years)} lots but "
                    f"{schedule_from} has {len(schedule)}. Every element is "
                    "priced against one shared schedule, so a lot an element "
                    "sits out is a zero rather than a missing row."
                )

            if kind == "amount":
                elements.append(wbs.flat_amount(el["name"], values))
                continue

            analogy_rows = [r for r in el["analogy"] if any(r)]
            if not analogy_rows:
                raise ValueError(
                    f"{el['name']} has no analogy lots. A fitted element "
                    "needs its own history to fit a curve to. If it has no "
                    "history, make it an amount or a factor instead."
                )
            elements.append(
                wbs.fitted(
                    el["name"],
                    self._analogy_frame(analogy_rows, el["name"]),
                    values,
                    cf,
                )
            )

        if not schedule:
            raise ValueError(
                "No element has a lot schedule. A programme of factors alone "
                "has nothing to be a percentage of."
            )
        return wbs.Program(
            name=self.var_program.get().strip() or "unnamed program",
            fiscal_years=schedule,
            elements=elements,
        )

    def _analogy_frame(self, rows, who):
        fy, qty, auc = [], [], []
        for i, r in enumerate(rows, start=1):
            try:
                q = parse_float(r[1])
            except (ValueError, IndexError):
                raise ValueError(
                    f"{who}, analogy row {i}: quantity {r[1]!r} is not a "
                    "number."
                )
            if q <= 0:
                raise ValueError(
                    f"{who}, analogy row {i}: quantity must be above zero."
                )
            fy.append(parse_float(r[0]) if r[0] else np.nan)
            qty.append(q)
            auc.append(parse_float(r[2]) if len(r) > 2 and r[2] else np.nan)
        return pd.DataFrame(
            {
                "Lot": range(1, len(rows) + 1),
                "Lot FY": fy,
                "Qty": qty,
                "AUC ($K)": auc,
            }
        )

    def _estimate_columns(self, rows, who):
        years, qty, cf = [], [], []
        try:
            last = float(self.var_defaultcf.get() or 1.0)
        except ValueError:
            last = 1.0
        for i, r in enumerate(rows, start=1):
            try:
                years.append(int(parse_float(r[0])))
            except (ValueError, IndexError):
                raise ValueError(
                    f"{who}, forecast row {i}: fiscal year {r[0]!r} is not a "
                    "number. Every lot needs one, because the schedule is "
                    "shared across elements."
                )
            # A blank quantity means this element is not bought in that lot.
            raw = r[1] if len(r) > 1 else ""
            try:
                q = parse_float(raw) if raw else 0.0
            except ValueError:
                raise ValueError(
                    f"{who}, forecast row {i}: quantity {raw!r} is not a "
                    "number."
                )
            qty.append(q)
            raw_cf = r[2] if len(r) > 2 else ""
            if raw_cf:
                last = parse_float(raw_cf)
            cf.append(last)
        return years, qty, cf

    def run_program(self):
        """Price every element and roll them up."""
        if wbs is None:
            return
        self.var_prog_status.set("Running...")
        self.update_idletasks()
        try:
            program = self.build_program()
            try:
                rho = float(self.var_prog_rho.get())
            except ValueError:
                raise ValueError("Correlation must be a number between 0 and 1.")

            result = wbs.roll_up(
                program,
                self._collect_overrides(),
                simulate=bool(self.var_prog_risk.get()),
                correlation=rho,
                n_iter=int(self.var_iters.get() or 20000)
                if hasattr(self, "var_iters") else 20000,
                seed=int(self.var_seed.get() or 11)
                if hasattr(self, "var_seed") else 11,
            )
        except (ValueError, wbs.ProgramError) as exc:
            messagebox.showerror("Program roll-up", str(exc))
            self.var_prog_status.set("Did not run.")
            return
        except Exception as exc:
            messagebox.showerror(
                "Program roll-up failed",
                f"{type(exc).__name__}: {exc}\n\n"
                f"{traceback.format_exc(limit=3)}",
            )
            self.var_prog_status.set("Failed.")
            return

        self.program_result = result
        self._show_program(result)
        self._show_program_extras(result)

        # The roll-up priced every element, so tab 4 can show the selected
        # one instead of saying no run has happened.
        selected = self.elements[self.current_element]["name"]
        mine = next(
            (e for e in result.elements if e.name == selected), None
        )
        if mine is not None:
            self._show_results(mine.summary, switch=False)
            self.lbl_result.config(
                text=(
                    f"{mine.name}, from the program roll-up. It selected "
                    f"{mine.model} and came to {mine.total:,.2f} before "
                    "risk. "
                    "Switch elements with the bar above the tabs; "
                    "tab 6 has the program."
                )
            )
        self.nb.select(self.tab_program)
        self.var_prog_status.set(
            f"{len(result.elements)} element(s). Total before risk "
            f"{result.total:,.2f}."
        )
        self._save_program_workbook(result)

    def _show_program(self, result):
        table = wbs.element_summary(result)
        self.tree_prog_elements.delete(
            *self.tree_prog_elements.get_children()
        )

        def cell(row, name, fmt="{:,.2f}"):
            if name not in row.index or row[name] == "" or pd.isna(row[name]):
                return ""
            try:
                return fmt.format(float(row[name]))
            except (TypeError, ValueError):
                return str(row[name])

        for _, r in table.iterrows():
            is_total = r["WBS Element"] == "PROGRAM TOTAL"
            self.tree_prog_elements.insert(
                "",
                "end",
                values=(
                    r["WBS Element"],
                    r["Model"],
                    cell(r, "Analogy lots", "{:.0f}"),
                    cell(r, "T1 ($K)"),
                    cell(r, "Units bought", "{:.0f}"),
                    cell(r, "Cost Before Risk ($)"),
                    cell(r, "Share of Program", "{:.1%}"),
                    cell(r, "P80 With Risk ($)"),
                ),
                tags=("total",) if is_total else (),
            )

        self.tree_prog_summary.delete(*self.tree_prog_summary.get_children())
        for _, r in wbs.program_summary(result).iterrows():
            item = str(r["Item"])
            self.tree_prog_summary.insert(
                "",
                "end",
                values=(item, str(r["Value"])),
                tags=("head",) if item.startswith("---") else (),
            )

    def _save_program_workbook(self, result):
        path = self.var_outfile.get().strip()
        if not path:
            return
        base, ext = os.path.splitext(path)
        target = f"{base}_program{ext or '.xlsx'}"
        while True:
            try:
                wbs.save_program_workbook(
                    target, result, getattr(self, 'sensitivity_result', None)
                )
                break
            except PermissionError:
                if not messagebox.askretrycancel(
                    "Cannot write the program workbook",
                    f"Could not write to:\n{target}\n\n"
                    "The file may be open in Excel.",
                ):
                    return
            except Exception as exc:
                messagebox.showwarning(
                    "Program workbook not saved",
                    f"The roll-up is on screen, but the workbook could not "
                    f"be written.\n\n{type(exc).__name__}: {exc}",
                )
                return
        self.var_prog_status.set(
            self.var_prog_status.get() + f"  Saved: {target}"
        )

    # -- WBS elements -------------------------------------------------------
    def _build_element_bar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=12, pady=(0, 6))
        ttk.Label(bar, text="WBS element:").pack(side="left")
        self.var_element = tk.StringVar()
        self.cmb_element = ttk.Combobox(
            bar, textvariable=self.var_element, state="readonly", width=34
        )
        self.cmb_element.pack(side="left", padx=6)
        self.cmb_element.bind("<<ComboboxSelected>>", self._on_element_change)
        ttk.Button(bar, text="Add", width=6,
                   command=self.add_element).pack(side="left")
        ttk.Button(bar, text="Rename", width=8,
                   command=self.rename_element).pack(side="left", padx=4)
        ttk.Button(bar, text="Remove", width=8,
                   command=self.remove_element).pack(side="left")
        self.var_element_hint = tk.StringVar(value="")
        ttk.Label(
            bar, textvariable=self.var_element_hint, style="Sub.TLabel"
        ).pack(side="left", padx=12)

    def _build_kind_bar(self, parent):
        """Kind-specific controls, shown only for the kind in hand."""
        self.kind_bar = ttk.Frame(parent)
        self.kind_bar.pack(fill="x", padx=12, pady=(0, 4))

        self.var_kind_note = tk.StringVar(value="")
        self.lbl_kind = ttk.Label(
            self.kind_bar, textvariable=self.var_kind_note, style="Sub.TLabel"
        )
        self.lbl_kind.pack(side="left")

        self.amount_box = ttk.Frame(self.kind_bar)
        ttk.Button(
            self.amount_box, text="Phase a total...",
            command=self.phase_amount,
        ).pack(side="left", padx=(6, 0))

        self.factor_box = ttk.Frame(self.kind_bar)
        ttk.Label(self.factor_box, text="Percentage:").pack(side="left")
        self.var_factor_pct = tk.StringVar(value="8.0")
        ent = ttk.Entry(self.factor_box, textvariable=self.var_factor_pct,
                        width=8)
        ent.pack(side="left", padx=(4, 2))
        ttk.Label(self.factor_box, text="%").pack(side="left")
        ttk.Button(
            self.factor_box, text="Choose what it applies to...",
            command=self.choose_basis,
        ).pack(side="left", padx=8)
        self.var_basis_note = tk.StringVar(value="")
        ttk.Label(
            self.factor_box, textvariable=self.var_basis_note,
            style="Sub.TLabel",
        ).pack(side="left")
        for var in (self.var_factor_pct,):
            var.trace_add("write", lambda *_: self._capture_element())

    def _refresh_element_views(self):
        """Point tabs 4 and 5 at the element now selected.

        They show one element at a time, so leaving the previous element's
        numbers up while the bar says something else is worse than showing
        nothing: the heading names one element and the table belongs to
        another.
        """
        if not self.elements:
            return
        el = self.elements[self.current_element]
        name, kind = el["name"], el.get("kind", "fitted")

        result = getattr(self, "program_result", None)
        mine = None
        if result is not None:
            mine = next(
                (e for e in result.elements if e.name == name), None
            )

        if mine is None:
            self.tree.delete(*self.tree.get_children())
            self.lbl_result.config(
                text=(
                    f"Nothing run for {name} yet. Click Run Model for this "
                    "element on its own, or run the program roll-up on tab 6."
                )
            )
        elif kind == "fitted":
            self._show_results(mine.summary, switch=False)
            self.lbl_result.config(
                text=(
                    f"{mine.name}, from the program roll-up. It selected "
                    f"{mine.model} and came to {mine.total:,.2f} before "
                    "risk. Switch elements with the bar above the tabs; "
                    "tab 6 has the program."
                )
            )
        else:
            # A factor or an amount has no curve, so there are no fit
            # statistics to show. Saying so beats an empty table.
            self.tree.delete(*self.tree.get_children())
            self.lbl_result.config(
                text=(
                    f"{mine.name} is {'a factor' if kind == 'factor' else 'an amount'} "
                    f"element, so it has no curve and no fit statistics. "
                    f"{mine.model}. It came to {mine.total:,.2f} before risk; "
                    "see tab 6 for it in the program."
                )
            )

        # The risk tab holds one element's intervals. Clear it rather than
        # let it be read as belonging to whichever element is now selected.
        owner = getattr(self, "risk_result_element", None)
        if owner is not None and owner != name:
            if getattr(self, "tree_risk", None) is not None:
                self.tree_risk.delete(*self.tree_risk.get_children())
                self.tree_iv.delete(*self.tree_iv.get_children())
            self.risk_result = None
            self.risk_result_element = None
            if hasattr(self, "var_risk_status"):
                self.var_risk_status.set(
                    f"Cleared: those intervals were for {owner}."
                )

    def _refresh_kind_bar(self):
        """Show the controls the current element's kind needs."""
        if not self.elements:
            return
        el = self.elements[self.current_element]
        kind = el.get("kind", "fitted")
        self.factor_box.pack_forget()
        self.amount_box.pack_forget()

        self._refresh_grids_for_kind(kind)

        if kind == "fitted":
            self.var_kind_note.set(
                "Fitted element: enter its analogy lots on tab 1 and its "
                "quantity per lot on tab 2."
            )
        elif kind == "amount":
            self.var_kind_note.set(
                "Amount element: no curve, no analogy history. Click Phase a "
                "total to enter one number and spread it, or type each "
                "year's cost on tab 2."
            )
            self.amount_box.pack(side="left")
        else:
            self.var_kind_note.set("Factor element: ")
            self.factor_box.pack(side="left")
            self.var_factor_pct.set(
                f"{float(el.get('factor', 0.08)) * 100:g}"
            )
            basis = el.get("basis") or []
            self.var_basis_note.set(
                "of every fitted element" if not basis
                else "of " + ", ".join(basis)
            )

    def _refresh_grids_for_kind(self, kind: str):
        """Point the two lot grids at what this kind of element actually is.

        The grids are shared, so without this an amount element shows a
        column headed Lot Quantity that the tool reads as dollars, and an
        analogy tab it never opens. Hide what does not apply so the heading
        on screen is always the number the engine will read.
        """
        if not hasattr(self, "grid_estimate"):
            return

        if kind == "fitted":
            self.grid_analogy.set_headers(
                ["Fiscal Year", "Lot Quantity", "Unit Cost AUC ($K)"]
            )
            self.grid_estimate.set_headers(
                ["Fiscal Year", "Lot Quantity", "Complexity Factor"]
            )
            wanted = {self.tab_analogy, self.tab_estimate}
        elif kind == "amount":
            # No curve and no history: a cost per year and nothing else.
            # Complexity multiplies a fitted curve, so it has no meaning on
            # a number that was quoted rather than estimated.
            self.grid_estimate.set_headers(["Fiscal Year", "Amount ($)", ""])
            wanted = {self.tab_estimate}
        else:
            # A factor is a percentage of other elements. It has no lots at
            # all, so both grids are noise.
            wanted = set()

        for tab in (self.tab_analogy, self.tab_estimate):
            self.nb.tab(tab, state=("normal" if tab in wanted else "hidden"))

        # Never leave the user staring at a tab that just disappeared.
        try:
            current = self.nb.nametowidget(self.nb.select())
        except (tk.TclError, KeyError):
            return
        if current in (self.tab_analogy, self.tab_estimate)                 and current not in wanted:
            self.nb.select(
                self.tab_estimate if self.tab_estimate in wanted
                else self.tab_run
            )

    def _first_tab_for_kind(self, kind: str):
        """The tab this kind of element is actually filled in on.

        Selecting a hidden tab makes Tk display it again, so anything that
        moves the user after a kind change has to pick a tab that kind still
        has.
        """
        if kind == "fitted":
            return self.tab_analogy
        if kind == "amount":
            return self.tab_estimate
        return self.tab_run

    def phase_amount(self):
        """Enter a non-recurring total and spread it over the lots.

        The per-lot amounts stay the thing that gets costed. This writes them
        for you, and every year is still editable afterwards, so the profile
        is a convenience rather than something the estimate depends on.
        """
        self._capture_element()
        el = self.elements[self.current_element]
        years = [r[0] for r in el["estimate"] if any(r)]
        if not years:
            # Borrow the schedule from an element that has one.
            for other in self.elements:
                other_years = [r[0] for r in other["estimate"] if any(r)]
                if other_years:
                    years = other_years
                    break
        if not years:
            messagebox.showinfo(
                "No schedule yet",
                "Enter the fiscal years on tab 2 first, or set them up on an "
                "element that has lots, so there is something to phase across.",
            )
            return

        win = tk.Toplevel(self)
        win.title(f"Phase a total across {len(years)} lots")
        win.transient(self)
        win.grab_set()

        ttk.Label(
            win,
            text=(
                "Non-recurring work is quoted as one number and then phased. "
                "Enter the total and how it\nfalls, and the lot amounts get "
                "filled in. You can edit any year afterwards."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(10, 8))

        row = ttk.Frame(win)
        row.pack(fill="x", padx=16)
        ttk.Label(row, text="Total ($):").pack(side="left")
        var_total = tk.StringVar(value="")
        ttk.Entry(row, textvariable=var_total, width=18).pack(
            side="left", padx=6
        )

        method = tk.StringVar(value="even")
        body = ttk.Frame(win)
        body.pack(fill="x", padx=16, pady=(8, 0))
        ttk.Radiobutton(
            body, text="Spread evenly across every lot",
            value="even", variable=method,
        ).pack(anchor="w")
        ttk.Radiobutton(
            body, text="All of it in one lot", value="single", variable=method
        ).pack(anchor="w")
        ttk.Radiobutton(
            body, text="By percentage, one per lot",
            value="percentages", variable=method,
        ).pack(anchor="w")

        detail = ttk.Frame(win)
        detail.pack(fill="x", padx=16, pady=(6, 0))
        ttk.Label(detail, text="Lot for 'one lot':").grid(
            row=0, column=0, sticky="e", pady=3
        )
        var_lot = tk.StringVar(value=years[0])
        ttk.Combobox(
            detail, textvariable=var_lot, values=years, state="readonly",
            width=10,
        ).grid(row=0, column=1, sticky="w", padx=6)
        ttk.Label(detail, text="Percentages:").grid(
            row=1, column=0, sticky="e", pady=3
        )
        var_pcts = tk.StringVar(
            value=", ".join(["0"] * len(years))
        )
        ttk.Entry(detail, textvariable=var_pcts, width=34).grid(
            row=1, column=1, sticky="w", padx=6
        )
        ttk.Label(
            detail,
            text=f"one per lot, in order: {', '.join(years)}; must total 100",
            style="Sub.TLabel",
        ).grid(row=2, column=1, sticky="w", padx=6)

        def apply():
            try:
                total = parse_float(var_total.get())
            except ValueError:
                messagebox.showerror(
                    "Phase a total", "The total must be a number.",
                    parent=win,
                )
                return
            how = method.get()
            try:
                if how == "single":
                    lots = [years.index(var_lot.get()) + 1]
                    amounts = wbs.phase_total(
                        total, len(years), "single", lots=lots
                    )
                elif how == "percentages":
                    pcts = [
                        float(x) for x in
                        var_pcts.get().replace(";", ",").split(",")
                        if x.strip()
                    ]
                    amounts = wbs.phase_total(
                        total, len(years), "percentages", percentages=pcts,
                        lots=list(range(1, len(pcts) + 1)),
                    )
                else:
                    amounts = wbs.phase_total(total, len(years), "even")
            except (ValueError, wbs.ProgramError) as exc:
                messagebox.showerror("Phase a total", str(exc), parent=win)
                return

            self.grid_estimate.load(
                [(y, f"{a:.2f}", "") for y, a in zip(years, amounts)]
            )
            self._capture_element()
            self.nb.select(self.tab_estimate)
            self.var_status.set(
                f"Phased {total:,.2f} across {len(years)} lots. Edit any year "
                "on tab 2 if the profile needs adjusting."
            )
            win.destroy()

        bar = ttk.Frame(win)
        bar.pack(fill="x", padx=12, pady=12)
        ttk.Button(bar, text="Apply", command=apply).pack(side="right")
        ttk.Button(bar, text="Cancel", command=win.destroy).pack(
            side="right", padx=6
        )
        win.wait_window()

    def choose_basis(self):
        """Pick which elements this factor is a percentage of."""
        el = self.elements[self.current_element]
        others = [
            e["name"] for e in self.elements if e["name"] != el["name"]
        ]
        if not others:
            messagebox.showinfo(
                "Nothing to apply to",
                "Add the elements this is a percentage of first.",
            )
            return

        win = tk.Toplevel(self)
        win.title(f"What is {el['name']} a percentage of?")
        win.transient(self)
        win.grab_set()
        ttk.Label(
            win,
            text=(
                "Tick the elements it applies to. Tick nothing to mean every "
                "fitted element,\nwhich is the usual reading of 'a percentage "
                "of the hardware'."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(10, 6))

        current = set(el.get("basis") or [])
        vars_ = {}
        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=16)
        for name in others:
            v = tk.BooleanVar(value=name in current)
            vars_[name] = v
            kind = next(
                (e.get("kind", "fitted") for e in self.elements
                 if e["name"] == name), "fitted"
            )
            ttk.Checkbutton(
                body, text=f"{name}   ({kind})", variable=v
            ).pack(anchor="w")

        def apply():
            chosen = [n for n, v in vars_.items() if v.get()]
            el["basis"] = chosen
            self._refresh_kind_bar()
            win.destroy()

        bar = ttk.Frame(win)
        bar.pack(fill="x", padx=12, pady=10)
        ttk.Button(bar, text="OK", command=apply).pack(side="right")
        ttk.Button(bar, text="Cancel", command=win.destroy).pack(
            side="right", padx=6
        )
        win.wait_window()

    def _ask_kind(self) -> str | None:
        """Which of the three shapes this new element takes."""
        win = tk.Toplevel(self)
        win.title("What kind of element?")
        win.transient(self)
        win.grab_set()
        choice = tk.StringVar(value="fitted")
        for value, label, blurb in (
            ("fitted", "Priced from its own lots",
             "Hardware with an analogy history. Gets its own learning curve."),
            ("factor", "A percentage of other elements",
             "Systems engineering, programme management. Scales with what it "
             "supports."),
            ("amount", "A cost entered lot by lot",
             "Tooling, qualification. Happens once and follows no curve."),
        ):
            ttk.Radiobutton(
                win, text=label, value=value, variable=choice
            ).pack(anchor="w", padx=16, pady=(10, 0))
            ttk.Label(win, text="      " + blurb, style="Sub.TLabel").pack(
                anchor="w", padx=16
            )
        out = {}

        def ok():
            out["kind"] = choice.get()
            win.destroy()

        bar = ttk.Frame(win)
        bar.pack(fill="x", padx=12, pady=12)
        ttk.Button(bar, text="OK", command=ok).pack(side="right")
        ttk.Button(bar, text="Cancel", command=win.destroy).pack(
            side="right", padx=6
        )
        win.wait_window()
        return out.get("kind")

    def _blank_element(self, name: str, kind: str = "fitted") -> dict:
        return {
            "name": name,
            "kind": kind,
            "analogy": [],
            "estimate": [],
            "factor": 0.08,
            "basis": [],
        }

    def _capture_element(self):
        """Copy what is on screen into the element it belongs to."""
        if self._switching or not self.elements:
            return
        el = self.elements[self.current_element]
        el["analogy"] = self.grid_analogy.get_rows()
        el["estimate"] = self.grid_estimate.get_rows()
        if el.get("kind") == "factor":
            try:
                el["factor"] = float(self.var_factor_pct.get()) / 100.0
            except (ValueError, AttributeError):
                pass

    def _show_element(self, index: int):
        self._switching = True
        try:
            el = self.elements[index]
            # Headings first: a row created under the wrong headings would
            # have the columns this kind does not use cleared underneath it.
            self._refresh_grids_for_kind(el.get("kind", "fitted"))
            self.grid_analogy.load(
                [tuple(r) for r in el["analogy"]] or [("", "", "")]
            )
            self.grid_estimate.load(
                [tuple(r) for r in el["estimate"]] or [("", "", "")]
            )
            self.current_element = index
            self.var_element.set(el["name"])
        finally:
            self._switching = False
        self._refresh_kind_bar()
        self._refresh_element_views()

    def _refresh_element_list(self, select: int | None = None):
        names = [e["name"] for e in self.elements]
        self.cmb_element["values"] = names
        index = self.current_element if select is None else select
        index = max(0, min(index, len(self.elements) - 1))
        self._show_element(index)
        self.var_element_hint.set(
            "one element; tab 6 rolls up the program"
            if len(self.elements) == 1
            else f"{len(self.elements)} elements; tab 6 rolls them up"
        )

    def _on_element_change(self, _event=None):
        target = self.var_element.get()
        for i, el in enumerate(self.elements):
            if el["name"] == target and i != self.current_element:
                self._capture_element()
                self._refresh_element_list(i)
                self.var_status.set(f"Editing {target}.")
                return

    def _unique_element_name(self, base: str) -> str:
        names = {e["name"] for e in self.elements}
        if base not in names:
            return base
        n = 2
        while f"{base} ({n})" in names:
            n += 1
        return f"{base} ({n})"

    def add_element(self):
        kind = self._ask_kind()
        if not kind:
            return
        name = self._ask_name("Add a WBS element", "Element name:")
        if not name:
            return
        self._capture_element()
        fresh = self._blank_element(self._unique_element_name(name), kind)
        # Carry the buy schedule across with no quantities, so the fiscal
        # years stay in step and only the counts have to be filled in.
        current = self.elements[self.current_element]["estimate"]
        fresh["estimate"] = [[row[0], "", ""] for row in current]
        self.elements.append(fresh)
        self._refresh_element_list(len(self.elements) - 1)
        self.nb.select(self._first_tab_for_kind(kind))
        hint = {
            "fitted": "Enter its analogy lots and its quantity for each lot.",
            "amount": "Click Phase a total to spread one number across the "
                      "years, or type each year's cost in the Amount column "
                      "on tab 2.",
            "factor": "Set its percentage and what it applies to, above the "
                      "tabs.",
        }[kind]
        self.var_status.set(f"Added {fresh['name']}. {hint}")

    def rename_element(self):
        if not self.elements:
            return
        current = self.elements[self.current_element]["name"]
        name = self._ask_name("Rename element", "New name:", current)
        if not name or name == current:
            return
        self.elements[self.current_element]["name"] = (
            self._unique_element_name(name)
        )
        self._capture_element()
        self._refresh_element_list(self.current_element)

    def remove_element(self):
        if len(self.elements) <= 1:
            messagebox.showinfo(
                "Cannot remove",
                "A program needs at least one element. Rename this one or "
                "clear its lots instead.",
            )
            return
        name = self.elements[self.current_element]["name"]
        if not messagebox.askyesno(
            "Remove element",
            f"Remove {name} and everything entered for it?",
        ):
            return
        self.elements.pop(self.current_element)
        self._refresh_element_list(max(0, self.current_element - 1))
        self.var_status.set(f"Removed {name}.")

    def _ask_name(self, title: str, prompt: str, initial: str = "") -> str:
        from tkinter import simpledialog

        value = simpledialog.askstring(
            title, prompt, initialvalue=initial, parent=self
        )
        return (value or "").strip()

    # -- saving and reloading a run -----------------------------------------
    def _build_menu(self):
        bar = tk.Menu(self)
        run_menu = tk.Menu(bar, tearoff=0)
        run_menu.add_command(
            label="Open Run...", accelerator="Ctrl+O", command=self.open_run
        )
        run_menu.add_command(
            label="Save Run", accelerator="Ctrl+S", command=self.save_run
        )
        run_menu.add_command(label="Save Run As...", command=self.save_run_as)
        run_menu.add_separator()
        run_menu.add_command(label="Load Example", command=self.load_example)
        bar.add_cascade(label="Run", menu=run_menu)
        self.config(menu=bar)
        self.bind_all("<Control-o>", lambda e: self.open_run())
        self.bind_all("<Control-s>", lambda e: self.save_run())

    def run_state(self) -> dict:
        """Everything needed to reproduce this run, as plain data."""
        self._capture_element()
        return {
            "format": RUN_FORMAT,
            "format_version": RUN_FORMAT_VERSION,
            "saved_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "tool_version": TOOL_VERSION,
            "run_info": self._run_info(),
            "settings": self._collect_overrides(),
            "risk": {
                "level": self.var_level.get(),
                "iterations": self.var_iters.get(),
                "seed": self.var_seed.get(),
                "lot_correlation": self.var_rho.get(),
                "run_with_model": bool(self.var_do_risk.get()),
            }
            if hasattr(self, "var_level")
            else {},
            # Kept for a reader expecting one element: the selected one.
            "analogy_lots": self.grid_analogy.get_rows(),
            "estimate_lots": self.grid_estimate.get_rows(),
            "elements": [
                {
                    "name": el["name"],
                    "kind": el.get("kind", "fitted"),
                    "analogy_lots": el["analogy"],
                    "estimate_lots": el["estimate"],
                    "factor": el.get("factor", 0.08),
                    "basis": list(el.get("basis") or []),
                }
                for el in self.elements
            ],
            "selected_element": self.current_element,
            "output_path": self.var_outfile.get().strip(),
        }

    def apply_run_state(self, data: dict):
        """Populate the window from a saved run.

        A file written before elements existed carries one set of lots, which
        loads as a single element rather than being rejected.
        """
        saved = data.get("elements")
        if saved:
            self.elements = [
                {
                    "name": e.get("name") or f"Element {i + 1}",
                    # A file written before the kinds existed is all fitted.
                    "kind": e.get("kind", "fitted"),
                    "analogy": [list(r) for r in e.get("analogy_lots", [])],
                    "estimate": [list(r) for r in e.get("estimate_lots", [])],
                    "factor": float(e.get("factor", 0.08)),
                    "basis": list(e.get("basis") or []),
                }
                for i, e in enumerate(saved)
            ]
        else:
            self.elements = [
                {
                    "name": "Element 1",
                    "kind": "fitted",
                    "analogy": [list(r) for r in data.get("analogy_lots", [])],
                    "estimate": [
                        list(r) for r in data.get("estimate_lots", [])
                    ],
                    "factor": 0.08,
                    "basis": [],
                }
            ]
        self.current_element = 0
        self._refresh_element_list(
            int(data.get("selected_element", 0) or 0)
        )

        info = data.get("run_info", {})
        self.var_runid.set(info.get("RunID", ""))
        self.var_program.set(info.get("Program", ""))
        self.var_label.set(info.get("RunLabel", ""))
        self.var_baseyear.set(info.get("BaseYear", ""))

        settings = data.get("settings", {})
        pairs = [
            ("CostUnitScale", self.var_costscale),
            ("TotalScale", self.var_totalscale),
            ("DefaultCF", self.var_defaultcf),
            ("TGate", self.var_tgate),
            ("FitPriorUnits", self.var_fitprior),
            ("FcstPriorUnits", self.var_fcstprior),
        ]
        for key, var in pairs:
            if key in settings:
                var.set(str(settings[key]))
        legacy = bool(settings.get("LegacyRateOmission", False))
        self.var_legacy_rate.set(legacy)

        risk_cfg = data.get("risk", {})
        for key, var in (
            ("level", getattr(self, "var_level", None)),
            ("iterations", getattr(self, "var_iters", None)),
            ("seed", getattr(self, "var_seed", None)),
            ("lot_correlation", getattr(self, "var_rho", None)),
        ):
            if var is not None and key in risk_cfg:
                var.set(str(risk_cfg[key]))
        if hasattr(self, "var_do_risk") and "run_with_model" in risk_cfg:
            self.var_do_risk.set(bool(risk_cfg["run_with_model"]))

        if data.get("output_path"):
            self.var_outfile.set(data["output_path"])

        if legacy:
            # Loading a legacy run silently would reproduce overstated costs
            # without saying so.
            messagebox.showwarning(
                "Legacy rate projection",
                "This run was saved with the legacy rate projection turned "
                "on, and that setting has been restored.\n\n"
                "Rate and LC+Rate costs will come out overstated. Untick it "
                "on the Run Info & Settings tab unless you are deliberately "
                "reproducing an old workbook.",
            )

    def open_run(self):
        path = filedialog.askopenfilename(
            title="Open a saved run",
            filetypes=[("Lot cost model run", "*.json"), ("All files", "*.*")],
            initialdir=os.path.dirname(self.run_path or "")
            or default_output_dir(),
        )
        if not path:
            return
        try:
            data = read_run_file(path)
        except RunFileError as exc:
            messagebox.showerror("Could not open run", str(exc))
            return
        try:
            self.apply_run_state(data)
        except Exception as exc:
            messagebox.showerror(
                "Could not load run",
                f"{type(exc).__name__}: {exc}\n\nThe file may be damaged.",
            )
            return
        self.run_path = path
        saved = data.get("saved_at", "")
        version = data.get("tool_version", "unknown")
        self.var_status.set(
            f"Loaded {os.path.basename(path)}"
            + (f" (saved {saved} by {version})" if saved else "")
        )

    def save_run(self):
        if not self.run_path:
            return self.save_run_as()
        return self._write_run(self.run_path)

    def save_run_as(self):
        base = self.var_program.get().strip() or "run"
        path = filedialog.asksaveasfilename(
            title="Save this run",
            defaultextension=".json",
            filetypes=[("Lot cost model run", "*.json")],
            initialfile=f"{base}{RUN_SUFFIX}",
            initialdir=os.path.dirname(self.run_path or "")
            or default_output_dir(),
        )
        if not path:
            return None
        return self._write_run(path)

    def _write_run(self, path: str):
        try:
            state = self.run_state()
        except ValueError as exc:
            messagebox.showerror("Check your settings", str(exc))
            return None
        while True:
            try:
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(state, fh, indent=2)
                break
            except PermissionError:
                if not messagebox.askretrycancel(
                    "Cannot write the file",
                    f"Could not write to:\n{path}\n\n"
                    "The file may be open elsewhere, or the folder may be "
                    "read-only.",
                ):
                    return None
        self.run_path = path
        self.var_status.set(f"Run saved: {path}")
        return path

    def load_example(self):
        kind = self.elements[self.current_element].get("kind", "fitted")
        if kind != "fitted":
            # The example is an analogy history and a quantity per lot.
            # Loaded onto an amount those quantities would be read as
            # dollars, which is the mistake the Amount heading exists to
            # prevent.
            messagebox.showinfo(
                "Example data",
                "The example is a learning curve, so it only fits a "
                f"fitted element. {self.elements[self.current_element]['name']}"
                f" is {'a factor' if kind == 'factor' else 'an amount'}. "
                "Switch to a fitted element, or add one, and load it there.",
            )
            return
        self.grid_analogy.load(EXAMPLE_ANALOGY)
        self.grid_estimate.load(EXAMPLE_ESTIMATE)
        self._capture_element()
        self.nb.select(self.tab_analogy)
        self.var_status.set(
            "Example data loaded into "
            f"{self.elements[self.current_element]['name']}."
        )

    def _build_risk(self):
        f = self.tab_risk

        available = risk is not None and risk.AVAILABLE
        if not available:
            why = (
                "risk.py is not next to this script."
                if risk is None
                else risk.IMPORT_ERROR
            )
            ttk.Label(
                f,
                text=(
                    "Risk analysis unavailable.\n\n"
                    + (risk.INSTALL_HINT if risk is not None else why)
                    + f"\n\nDetail: {why}"
                ),
                style="Sub.TLabel",
                justify="left",
            ).pack(anchor="w", padx=12, pady=14)
            self.var_do_risk = tk.BooleanVar(value=False)
            self.tree_risk = None
            self.tree_iv = None
            return

        ttk.Label(
            f,
            text=(
                "Prediction intervals and a Monte Carlo of the whole buy, "
                "from cost_core in the cost-risk-toolkit.\nNothing is refitted: "
                "these put a distribution around the model already selected on "
                "tab 4, so the\npoint estimate underneath is the same buy total "
                "the projections sheet reports."
            ),
            style="Sub.TLabel",
            justify="left",
        ).pack(anchor="w", padx=8, pady=(8, 4))

        ctl = ttk.LabelFrame(f, text="Settings")
        ctl.pack(fill="x", padx=8, pady=4)

        self.var_level = tk.StringVar(value="80")
        self.var_iters = tk.StringVar(value="20000")
        self.var_seed = tk.StringVar(value="11")
        self.var_rho = tk.StringVar(value="0.30")
        self.var_do_risk = tk.BooleanVar(value=True)

        def combo(row, col, label, var, values, width=12):
            ttk.Label(ctl, text=label + ":").grid(
                row=row, column=col * 2, sticky="e", padx=(8, 4), pady=4
            )
            w = ttk.Combobox(
                ctl, textvariable=var, values=values, width=width,
                state="readonly",
            )
            w.grid(row=row, column=col * 2 + 1, sticky="w", pady=4)
            return w

        def entry(row, col, label, var, width=12):
            ttk.Label(ctl, text=label + ":").grid(
                row=row, column=col * 2, sticky="e", padx=(8, 4), pady=4
            )
            ttk.Entry(ctl, textvariable=var, width=width).grid(
                row=row, column=col * 2 + 1, sticky="w", pady=4
            )

        combo(0, 0, "Interval %", self.var_level, ["80", "90", "95"], 8)
        entry(0, 1, "Iterations", self.var_iters)
        entry(0, 2, "Seed", self.var_seed)
        entry(1, 0, "Lot correlation", self.var_rho, 8)
        ttk.Label(
            ctl,
            text="how much consecutive lots move together",
            style="Sub.TLabel",
        ).grid(row=1, column=2, columnspan=4, sticky="w", padx=8)
        ttk.Checkbutton(
            ctl,
            text="Run this automatically with Run Model and add it to the workbook",
            variable=self.var_do_risk,
        ).grid(row=2, column=0, columnspan=6, sticky="w", padx=8, pady=(2, 6))

        bar = ttk.Frame(f)
        bar.pack(fill="x", padx=8, pady=(2, 6))
        ttk.Button(
            bar, text="Run Risk Analysis", command=self.run_risk
        ).pack(side="left")
        self.var_risk_status = tk.StringVar(
            value="Not run yet. Needs a base year on tab 3."
        )
        ttk.Label(
            bar, textvariable=self.var_risk_status, style="Sub.TLabel"
        ).pack(side="left", padx=10)

        panes = ttk.Panedwindow(f, orient="vertical")
        panes.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        top = ttk.Frame(panes)
        self.tree_risk = ttk.Treeview(
            top, columns=("Item", "Value"), show="headings", height=9
        )
        self.tree_risk.heading("Item", text="Item")
        self.tree_risk.heading("Value", text="Value")
        self.tree_risk.column("Item", width=230, anchor="w")
        self.tree_risk.column("Value", width=680, anchor="w")
        sv = ttk.Scrollbar(top, orient="vertical",
                           command=self.tree_risk.yview)
        self.tree_risk.configure(yscrollcommand=sv.set)
        sv.pack(side="right", fill="y")
        self.tree_risk.pack(fill="both", expand=True)
        panes.add(top, weight=3)

        bot = ttk.Frame(panes)
        iv_cols = (
            "Lot", "Qty", "FY",
            "Unit Cost ($K)", "Unit Low", "Unit High",
            "Lot Cost ($)", "Lot Low", "Lot High",
        )
        self.tree_iv = ttk.Treeview(
            bot, columns=iv_cols, show="headings", height=8
        )
        for c, w in zip(iv_cols, (45, 45, 55, 105, 95, 95, 125, 115, 115)):
            self.tree_iv.heading(c, text=c)
            self.tree_iv.column(c, width=w, anchor="e")
        sv2 = ttk.Scrollbar(bot, orient="vertical", command=self.tree_iv.yview)
        self.tree_iv.configure(yscrollcommand=sv2.set)
        sv2.pack(side="right", fill="y")
        self.tree_iv.pack(fill="both", expand=True)
        panes.add(bot, weight=2)

    def _risk_options(self) -> "risk.RiskOptions":
        def num(var, label, caster):
            try:
                return caster(var.get().strip())
            except ValueError:
                raise ValueError(f"Risk setting '{label}' must be a number.")

        return risk.RiskOptions(
            level=num(self.var_level, "Interval %", float) / 100.0,
            n_iter=num(self.var_iters, "Iterations", int),
            seed=num(self.var_seed, "Seed", int),
            lot_correlation=num(self.var_rho, "Lot correlation", float),
            simulate=True,
        )

    def _compute_risk(self, ctx=None, projections=None, summary=None):
        """Put intervals and a simulated total around a fitted run.

        Given nothing, it fits from the grids first. Run Model passes in what
        it already computed so the model is not fitted twice.
        """
        if ctx is None or projections is None or summary is None:
            analogy_df = self._collect_analogy()
            estimate_df = self._collect_estimate()
            projections, ctx = run_lot_cost_model(
                analogy_df, estimate_df, self._collect_overrides()
            )
            summary = generate_analyst_summary(ctx, self._run_info())

        return risk.run_risk(ctx, projections, summary, self._risk_options())

    def _run_info(self) -> dict:
        return {
            "RunID": self.var_runid.get().strip() or SETTINGS["DefaultRunID"],
            "Program": self.var_program.get().strip()
            or SETTINGS["DefaultProgram"],
            "RunLabel": self.var_label.get().strip()
            or SETTINGS["DefaultRunLabel"],
            "BaseYear": self.var_baseyear.get().strip(),
        }

    def run_risk(self):
        if risk is None or not risk.AVAILABLE:
            messagebox.showinfo("Risk analysis unavailable", risk.INSTALL_HINT)
            return
        self.var_risk_status.set("Running...")
        self.update_idletasks()
        try:
            self.risk_result = self._compute_risk()
            self.risk_result_element = (
                self.elements[self.current_element]["name"]
            )
            self._show_risk(self.risk_result)
            self.var_risk_status.set(
                f"Ran {self.risk_result.n_iter:,} iterations."
                if self.risk_result.n_iter
                else "Intervals computed."
            )
            self.nb.select(self.tab_risk)
        except (ValueError, RuntimeError) as exc:
            messagebox.showerror("Risk analysis", str(exc))
            self.var_risk_status.set("Did not run.")
        except Exception as exc:
            messagebox.showerror(
                "Risk analysis failed",
                f"{type(exc).__name__}: {exc}\n\n"
                f"{traceback.format_exc(limit=3)}",
            )
            self.var_risk_status.set("Failed.")

    def _show_risk(self, res):
        self.tree_risk.delete(*self.tree_risk.get_children())
        for _, row in risk.summary_frame(res).iterrows():
            self.tree_risk.insert(
                "", "end", values=(str(row["Item"]), str(row["Value"]))
            )
        self.tree_iv.delete(*self.tree_iv.get_children())

        def cell(row, name, fmt="{:,.2f}"):
            # Lot and Fiscal Year come through as text, so anything that is
            # not a number is shown as it stands rather than formatted.
            if name not in row.index or pd.isna(row[name]):
                return ""
            value = row[name]
            try:
                return fmt.format(float(value))
            except (TypeError, ValueError):
                return str(value)

        for _, r in res.intervals.iterrows():
            self.tree_iv.insert(
                "",
                "end",
                values=(
                    cell(r, "Lot", "{:.0f}"),
                    cell(r, "Lot Quantity", "{:.0f}"),
                    cell(r, "Fiscal Year", "{:.0f}"),
                    cell(r, "Unit Cost ($K)"),
                    cell(r, "Unit Cost Lower"),
                    cell(r, "Unit Cost Upper"),
                    cell(r, "Lot Cost ($)"),
                    cell(r, "Lot Cost Lower"),
                    cell(r, "Lot Cost Upper"),
                ),
            )

    def _build_actionbar(self):
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=12, pady=10)

        ttk.Label(bar, text="Save workbook to:").pack(side="left")
        self.var_outfile = tk.StringVar(
            value=os.path.join(
                default_output_dir(), "Lot_Cost_Model_Complete_Suite.xlsx"
            )
        )
        ttk.Entry(bar, textvariable=self.var_outfile).pack(
            side="left", fill="x", expand=True, padx=6
        )
        ttk.Button(bar, text="Browse...", command=self._browse).pack(
            side="left"
        )
        self.btn_run = ttk.Button(
            bar, text="Run Model", command=self.run_model
        )
        self.btn_run.pack(side="left", padx=(10, 0))

        self.var_status = tk.StringVar(value="Ready.")
        ttk.Label(
            self, textvariable=self.var_status, style="Sub.TLabel"
        ).pack(anchor="w", padx=12, pady=(0, 8))

    def _browse(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
            initialfile=os.path.basename(self.var_outfile.get()),
            initialdir=os.path.dirname(self.var_outfile.get())
            or default_output_dir(),
            title="Save model workbook as",
        )
        if path:
            self.var_outfile.set(path)

    # -- input parsing ------------------------------------------------------
    def _collect_analogy(self) -> pd.DataFrame:
        rows = self.grid_analogy.get_rows()
        if not rows:
            raise ValueError("No analogy lots entered (tab 1).")

        fy, qty, auc = [], [], []
        for i, r in enumerate(rows, start=1):
            try:
                q = parse_float(r[1])
            except ValueError:
                raise ValueError(
                    f"Analogy row {i}: Lot Quantity '{r[1]}' is not a number."
                )
            if q <= 0:
                raise ValueError(
                    f"Analogy row {i}: Lot Quantity must be greater than 0."
                )
            try:
                y = parse_float(r[0]) if r[0] else np.nan
            except ValueError:
                raise ValueError(
                    f"Analogy row {i}: Fiscal Year '{r[0]}' is not a number."
                )
            if r[2]:
                try:
                    c = parse_float(r[2])
                except ValueError:
                    raise ValueError(
                        f"Analogy row {i}: Unit Cost '{r[2]}' is not a number."
                    )
                if c <= 0:
                    raise ValueError(
                        f"Analogy row {i}: Unit Cost must be greater than 0 "
                        "(leave it blank for a quantity-only lot)."
                    )
            else:
                c = np.nan
            fy.append(y)
            qty.append(q)
            auc.append(c)

        n_costed = sum(1 for c in auc if pd.notna(c))
        if n_costed < 3:
            raise ValueError(
                "The learning curve needs at least 3 analogy lots with both a "
                f"quantity and a unit cost. Found {n_costed}."
            )

        return pd.DataFrame(
            {
                "Lot": list(range(1, len(rows) + 1)),
                "Lot FY": fy,
                "Qty": qty,
                "AUC ($K)": auc,
            }
        )

    def _collect_estimate(self) -> pd.DataFrame:
        rows = self.grid_estimate.get_rows()
        if not rows:
            raise ValueError("No estimate lots entered (tab 2).")

        fy, qty, cf = [], [], []
        for i, r in enumerate(rows, start=1):
            try:
                q = parse_float(r[1])
            except ValueError:
                raise ValueError(
                    f"Estimate row {i}: Lot Quantity '{r[1]}' is not a number."
                )
            if q <= 0:
                raise ValueError(
                    f"Estimate row {i}: Lot Quantity must be greater than 0."
                )
            try:
                y = parse_float(r[0]) if r[0] else np.nan
            except ValueError:
                raise ValueError(
                    f"Estimate row {i}: Fiscal Year '{r[0]}' is not a number."
                )
            if r[2]:
                try:
                    c = parse_float(r[2])
                except ValueError:
                    raise ValueError(
                        f"Estimate row {i}: Complexity Factor '{r[2]}' is not "
                        "a number."
                    )
            else:
                c = np.nan
            fy.append(y)
            qty.append(q)
            cf.append(c)

        return pd.DataFrame(
            {
                "Lot": list(range(1, len(rows) + 1)),
                "Lot FY": fy,
                "Qty": qty,
                "Complexity": cf,
            }
        )

    def _collect_overrides(self) -> dict:
        def num(var, label, caster=float):
            try:
                return caster(var.get().strip())
            except ValueError:
                raise ValueError(f"Setting '{label}' must be a number.")

        return {
            "CostUnitScale": num(self.var_costscale, "Cost unit scale"),
            "TotalScale": num(self.var_totalscale, "Total scale"),
            "DefaultCF": num(self.var_defaultcf, "Default complexity"),
            "TGate": num(self.var_tgate, "t-gate"),
            "FitPriorUnits": num(self.var_fitprior, "Prior units (fit)", int),
            "FcstPriorUnits": num(
                self.var_fcstprior, "Prior units (forecast)", int
            ),
            "LegacyRateOmission": bool(self.var_legacy_rate.get()),
        }

    def _save_workbook(
        self,
        path,
        proj,
        summ,
        chart,
        risk_summ=None,
        risk_iv=None,
        risk_sc=None,
    ) -> str | None:
        """Save, re-prompting if the path is locked or not writable."""
        while True:
            try:
                save_complete_excel_workbook(
                    path, proj, summ, chart, risk_summ, risk_iv, risk_sc
                )
                return path
            except PermissionError:
                retry = messagebox.askretrycancel(
                    "Cannot write the file",
                    f"Could not write to:\n{path}\n\n"
                    "The file may be open in Excel, or the folder may be "
                    "read-only.\n\nClose the file and click Retry, or Cancel "
                    "to choose a different location.",
                )
                if retry:
                    continue
                new = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel workbook", "*.xlsx")],
                    initialfile=os.path.basename(path),
                    initialdir=default_output_dir(),
                    title="Save model workbook as",
                )
                if not new:
                    return None
                path = new
                self.var_outfile.set(path)

    # -- run ----------------------------------------------------------------
    def run_model(self):
        self.btn_run.config(state="disabled")
        self.var_status.set("Running...")
        self.update_idletasks()
        try:
            analogy_df = self._collect_analogy()
            estimate_df = self._collect_estimate()
            overrides = self._collect_overrides()

            run_info = self._run_info()

            projections_df, models_ctx = run_lot_cost_model(
                analogy_df, estimate_df, overrides
            )
            summary_df = generate_analyst_summary(models_ctx, run_info)
            chart_df = generate_fit_chart_data(models_ctx)

            risk_summary = risk_intervals = risk_scurve = None
            if (
                risk is not None
                and risk.AVAILABLE
                and bool(self.var_do_risk.get())
            ):
                try:
                    self.risk_result = self._compute_risk(
                        models_ctx, projections_df, summary_df
                    )
                    self.risk_result_element = (
                        self.elements[self.current_element]["name"]
                    )
                    self._show_risk(self.risk_result)
                    risk_summary = risk.summary_frame(self.risk_result)
                    risk_intervals = self.risk_result.intervals
                    risk_scurve = self.risk_result.scurve
                    self.var_risk_status.set(
                        f"Ran {self.risk_result.n_iter:,} iterations."
                    )
                except (ValueError, RuntimeError) as exc:
                    # A risk failure must not cost the analyst the main run.
                    self.var_risk_status.set("Did not run.")
                    messagebox.showwarning(
                        "Risk analysis skipped",
                        f"{exc}\n\nThe rest of the model ran normally.",
                    )

            path = self.var_outfile.get().strip()
            if not path:
                raise ValueError("Choose an output file first.")
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            saved = self._save_workbook(
                path,
                projections_df,
                summary_df,
                chart_df,
                risk_summary,
                risk_intervals,
                risk_scurve,
            )

            self._show_results(summary_df)
            if saved:
                self.var_outfile.set(saved)
                self.var_status.set(f"Saved: {saved}")
                self.lbl_result.config(
                    text=(
                        f"Run complete. Workbook saved to:\n{saved}\n"
                        "Sheets: Analyst_Summary, Estimate_Projections, "
                        "Fit_Chart_Data (with 3 embedded charts)."
                    )
                )
                if messagebox.askyesno(
                    "Run complete",
                    f"Model ran successfully.\n\nSaved to:\n{saved}\n\n"
                    "Open the workbook now?",
                ):
                    try:
                        os.startfile(saved)
                    except Exception:
                        pass
            else:
                self.var_status.set("Run complete, workbook not saved.")
                self.lbl_result.config(
                    text="Run complete, but the workbook was not saved."
                )

        except ValueError as exc:
            messagebox.showerror("Check your input", str(exc))
            self.var_status.set("Input error.")
        except Exception as exc:
            messagebox.showerror(
                "Model error",
                f"{type(exc).__name__}: {exc}\n\n"
                f"{traceback.format_exc(limit=3)}",
            )
            self.var_status.set("Run failed.")
        finally:
            self.btn_run.config(state="normal")

    def _show_results(self, summary_df: pd.DataFrame, switch: bool = True):
        self.tree.delete(*self.tree.get_children())
        for _, row in summary_df.iterrows():
            vals = [
                str(row["Item"]),
                str(row["Value"]),
                str(row["LC"]),
                str(row["Rate"]),
                str(row["LC+Rate"]),
            ]
            tag = "sel" if vals[0] == "SELECTED" else ""
            self.tree.insert("", "end", values=vals, tags=(tag,))
        if switch:
            self.nb.select(self.tab_results)


# ============================================================================
# 8. EXECUTION BLOCK
# ============================================================================
def main() -> int:
    """Open the window. Importable, so the tool can be packaged or scripted."""
    # Crisper text on high-DPI Windows displays.
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass

    try:
        app = LotCostApp()
    except tk.TclError as exc:
        print(f"Could not start the GUI: {exc}", file=sys.stderr)
        print(
            "A desktop session is required to run this tool.", file=sys.stderr
        )
        return 1

    app.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(main())
