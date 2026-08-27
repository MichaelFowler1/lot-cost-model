"""Roll several WBS elements into one program estimate.

Each element carries its own analogy history and gets its own curve: the
airframe learns at one rate, the engines at another, and neither is told
about the other. What they share is the buy schedule. The fiscal years are
a property of the programme, while the quantity per lot belongs to the
element, because a kit buy or a spares provision changes how many units of
that element you buy without changing when you buy them.

Rolling the point estimates up is arithmetic. Rolling the *risk* up is not,
and it is the reason this module exists rather than a column of SUMs in a
spreadsheet. Elements on one programme share a workforce, a supply base and
a schedule, so their overruns arrive together. Adding independent
distributions understates the variance of the total by a factor of
1 + rho(k-1), which for ten elements at rho = 0.3 is 3.7 in variance, and
the error lands on the upper tail where the P80 lives. The correlated
roll-up here is cost_core's, which exists for exactly this.
"""

from __future__ import annotations

from dataclasses import dataclass, field, replace

import numpy as np
import pandas as pd

from lot_cost_model import (
    SETTINGS,
    generate_analyst_summary,
    run_lot_cost_model,
)

try:
    from cost_core.lots import (
        influence_diagnostics,
        selected_model_name,
        simulate_buy,
    )
    from cost_core.monte_carlo import (
        CostElement,
        RiskModel,
        correlation_impact,
        simulate_risk_model,
    )

    RISK_AVAILABLE = True
    RISK_IMPORT_ERROR = ""
except Exception as exc:  # pragma: no cover - depends on the environment
    RISK_AVAILABLE = False
    RISK_IMPORT_ERROR = f"{type(exc).__name__}: {exc}"

#: cost_core's own default. Stated rather than hidden: it is an assumption,
#: and an unstated assumption is the failure this whole module guards against.
DEFAULT_CORRELATION = 0.25


class ProgramError(Exception):
    """The programme as described cannot be priced."""


#: The three shapes a WBS element can take.
#:
#: ``fitted``  Priced from its own analogy history by a learning curve.
#: ``factor``  A percentage of other elements. Systems engineering and
#:             programme management are the usual cases: they do not learn,
#:             they scale with the hardware they support.
#: ``amount``  A cost entered directly, lot by lot. Tooling, qualification,
#:             and anything else that happens once and follows no curve.
KINDS = ("fitted", "factor", "amount")


@dataclass
class Element:
    """One WBS element.

    Only ``fitted`` elements carry a curve. Forcing a learning curve onto
    systems engineering or a one-off tooling buy would be fitting a shape the
    work does not have, so the other two kinds are priced arithmetically and
    say so.
    """

    name: str
    kind: str = "fitted"

    # -- fitted --------------------------------------------------------------
    #: Analogy lots for this element alone. Columns: Lot FY, Qty, AUC ($K).
    analogy: pd.DataFrame | None = None
    #: Units of this element bought in each lot of the programme schedule.
    quantities: list[float] | None = None
    #: Complexity factor per lot. A single value is broadcast.
    complexity: list[float] | float = 1.0
    #: Per-element overrides, merged over SETTINGS.
    settings: dict = field(default_factory=dict)

    # -- factor --------------------------------------------------------------
    #: Share of the basis, as a fraction. 0.08 is eight percent.
    factor: float | None = None
    #: Elements it is a percentage of. Empty or None means every fitted
    #: element, which is the usual reading of "a percentage of hardware".
    basis: list[str] | None = None

    # -- amount --------------------------------------------------------------
    #: Cost in each lot of the programme schedule, already in the same
    #: dollars as everything else. Inflation is assumed already applied.
    amounts: list[float] | None = None

    def complexity_per_lot(self, n_lots: int) -> list[float]:
        if isinstance(self.complexity, (int, float)):
            return [float(self.complexity)] * n_lots
        return [float(c) for c in self.complexity]

    def validate(self, n_lots: int):
        """Check this element carries what its kind needs."""
        if self.kind not in KINDS:
            raise ProgramError(
                f"{self.name} has kind {self.kind!r}; expected one of "
                f"{list(KINDS)}."
            )
        if self.kind == "fitted":
            if self.analogy is None or not len(self.analogy):
                raise ProgramError(
                    f"{self.name} is a fitted element, so it needs analogy "
                    "lots of its own to fit a curve to."
                )
            if self.quantities is None:
                raise ProgramError(
                    f"{self.name} is a fitted element and needs a quantity "
                    "for each lot."
                )
            if len(self.quantities) != n_lots:
                raise ProgramError(
                    f"{self.name} has {len(self.quantities)} lot quantities "
                    f"but the programme has {n_lots} lots. Every element is "
                    "priced against the same schedule; a lot it does not "
                    "take part in is a quantity of zero."
                )
            if any(q < 0 for q in self.quantities):
                raise ProgramError(
                    f"{self.name} has a negative lot quantity."
                )
            if not any(q > 0 for q in self.quantities):
                raise ProgramError(
                    f"{self.name} is bought in no lot, so it has no cost. "
                    "Remove it or give it a quantity."
                )
        elif self.kind == "factor":
            if self.factor is None:
                raise ProgramError(
                    f"{self.name} is a factor element and needs a factor, "
                    "as a fraction: 0.08 for eight percent."
                )
            if self.factor < 0:
                raise ProgramError(
                    f"{self.name} has a negative factor."
                )
        else:  # amount
            if self.amounts is None:
                raise ProgramError(
                    f"{self.name} is an amount element and needs a cost for "
                    "each lot; use zero for the lots it does not fall in."
                )
            if len(self.amounts) != n_lots:
                raise ProgramError(
                    f"{self.name} has {len(self.amounts)} lot amounts but "
                    f"the programme has {n_lots} lots."
                )
            if any(a < 0 for a in self.amounts):
                raise ProgramError(f"{self.name} has a negative amount.")


def fitted(name, analogy, quantities, complexity=1.0, **kw) -> Element:
    """A hardware element priced from its own history."""
    return Element(
        name=name, kind="fitted", analogy=analogy, quantities=quantities,
        complexity=complexity, **kw
    )


def factor_of(name, factor, basis=None) -> Element:
    """A percentage of other elements, such as SE/PM on the hardware."""
    return Element(name=name, kind="factor", factor=factor, basis=basis)


def flat_amount(name, amounts) -> Element:
    """A cost entered lot by lot, such as tooling or qualification."""
    return Element(name=name, kind="amount", amounts=list(amounts))


@dataclass
class Program:
    """A shared buy schedule and the elements bought against it."""

    name: str
    #: One fiscal year per lot. Every element is priced against these lots.
    fiscal_years: list[int]
    elements: list[Element] = field(default_factory=list)

    @property
    def n_lots(self) -> int:
        return len(self.fiscal_years)

    def validate(self):
        if not self.fiscal_years:
            raise ProgramError("The programme needs at least one lot.")
        if not self.elements:
            raise ProgramError("The programme needs at least one element.")
        seen = set()
        for el in self.elements:
            if el.name in seen:
                raise ProgramError(f"Two elements are both named {el.name!r}.")
            seen.add(el.name)
            el.validate(self.n_lots)

        if not any(e.kind == "fitted" for e in self.elements):
            raise ProgramError(
                "The programme has no fitted element. A factor needs "
                "something to be a percentage of, and an amount on its own "
                "is a number, not an estimate."
            )
        # Resolving here rather than at pricing time so a bad basis or a
        # circular reference is reported before anything is computed.
        self.pricing_order()

    def pricing_order(self) -> list[Element]:
        """Elements in an order where every basis is priced before its factor.

        Raises:
            ProgramError: On a basis naming an element that does not exist,
                an element used as its own basis, or a cycle between factors.
        """
        by_name = {e.name: e for e in self.elements}
        ordered: list[Element] = [
            e for e in self.elements if e.kind in ("fitted", "amount")
        ]
        done = {e.name for e in ordered}
        pending = [e for e in self.elements if e.kind == "factor"]

        for el in pending:
            for ref in el.basis or []:
                if ref not in by_name:
                    raise ProgramError(
                        f"{el.name} is a percentage of {ref!r}, which is not "
                        f"an element of this programme. Known elements: "
                        f"{sorted(by_name)}."
                    )
                if ref == el.name:
                    raise ProgramError(
                        f"{el.name} is a percentage of itself."
                    )

        while pending:
            ready = [
                e for e in pending
                if all(ref in done for ref in (e.basis or []))
            ]
            if not ready:
                stuck = ", ".join(sorted(e.name for e in pending))
                raise ProgramError(
                    "These factor elements depend on each other in a circle, "
                    f"so none can be priced first: {stuck}."
                )
            for e in ready:
                ordered.append(e)
                done.add(e.name)
            pending = [e for e in pending if e not in ready]
        return ordered


@dataclass
class ElementResult:
    """What one element came to, and how."""

    name: str
    model: str
    projections: pd.DataFrame
    summary: pd.DataFrame
    ctx: dict
    total: float
    by_lot: np.ndarray
    n_lots_fitted: int
    kind: str = "fitted"
    #: For a factor element, what it was a percentage of.
    basis: list[str] = field(default_factory=list)
    #: Present only when the element was simulated.
    totals: np.ndarray | None = None


@dataclass
class ProgramResult:
    """The rolled-up estimate, with every element still visible inside it."""

    program: str
    elements: list[ElementResult]
    by_lot: pd.DataFrame
    total: float
    #: Correlated programme distribution, when the risk roll-up ran.
    p50: float | None = None
    p80: float | None = None
    p90: float | None = None
    mean: float | None = None
    std: float | None = None
    cv: float | None = None
    point_percentile: float | None = None
    correlation: float | None = None
    n_iter: int | None = None
    scurve: pd.DataFrame | None = None
    #: Variance contribution by element, largest first.
    tornado: pd.DataFrame | None = None
    #: Elements in the order they were priced, bases before their factors.
    order: list = field(default_factory=list)
    #: What treating the elements as independent would have understated.
    independence_understates_sd_by: float | None = None
    #: The closed-form 1 + rho(k-1), to check the sampler against the algebra.
    variance_ratio_analytic: float | None = None
    #: How much lower the P80 would sit under independence, as a fraction.
    p80_understatement: float | None = None
    #: The share of the P80 reserve an independence assumption throws away.
    reserve_understatement: float | None = None
    notes: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _estimate_frame(program: Program, element: Element) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Lot": range(1, program.n_lots + 1),
            "Lot FY": program.fiscal_years,
            "Qty": [float(q) for q in element.quantities],
            "Complexity": element.complexity_per_lot(program.n_lots),
        }
    )


def price_derived(
    program: Program, element: Element, priced: "dict[str, ElementResult]"
) -> ElementResult:
    """Cost a factor or amount element, which have no curve of their own.

    A factor is applied lot by lot rather than to the total, so it inherits
    the phasing of whatever it supports: engineering effort follows the
    hardware it is engineering.
    """
    n = program.n_lots
    if element.kind == "amount":
        by_lot = np.asarray(element.amounts, dtype=float)
        model = "Amount"
        basis: list[str] = []
    else:
        basis = list(element.basis or [
            name for name, r in priced.items() if r.kind == "fitted"
        ])
        if not basis:
            raise ProgramError(
                f"{element.name} is a percentage of nothing. Name the "
                "elements it applies to, or add a fitted element."
            )
        missing = [b for b in basis if b not in priced]
        if missing:
            raise ProgramError(
                f"{element.name} is a percentage of {missing}, which have "
                "not been priced."
            )
        base = np.sum([priced[b].by_lot for b in basis], axis=0)
        by_lot = float(element.factor) * base
        model = f"{element.factor:.1%} of " + ", ".join(basis)

    empty = pd.DataFrame(
        {
            "Lot": range(1, n + 1),
            "Fiscal Year": program.fiscal_years,
            "Lot Quantity": [0] * n,
            "Cost ($)": np.round(by_lot, 2),
        }
    )
    return ElementResult(
        name=element.name,
        model=model,
        projections=empty,
        summary=pd.DataFrame(columns=["Item", "Value", "LC", "Rate",
                                      "LC+Rate"]),
        ctx={},
        total=float(by_lot.sum()),
        by_lot=by_lot,
        n_lots_fitted=0,
        kind=element.kind,
        basis=basis,
    )


def price_element(
    program: Program, element: Element, overrides: dict | None = None
) -> ElementResult:
    """Fit this element's own curve and price its share of the buy.

    A lot the element is not bought in carries a quantity of zero, which the
    engine cannot price, so those lots are dropped for the fit and added back
    as zero cost afterwards. That keeps every element's by-lot vector the same
    length as the programme schedule.
    """
    estimate = _estimate_frame(program, element)
    priced_mask = estimate["Qty"].to_numpy(dtype=float) > 0
    if not priced_mask.any():
        raise ProgramError(f"{element.name} is bought in no lot.")

    cfg = dict(overrides or {})
    cfg.update(element.settings)

    try:
        projections, ctx = run_lot_cost_model(
            element.analogy, estimate.loc[priced_mask].copy(), cfg or None
        )
    except ValueError as exc:
        raise ProgramError(f"{element.name}: {exc}") from exc

    summary = generate_analyst_summary(ctx, {"Program": element.name})
    model = _selected_model(summary)

    cost_col = f"{model} Lot Cost After Complexity ($)"
    priced = projections[cost_col].to_numpy(dtype=float)

    by_lot = np.zeros(program.n_lots, dtype=float)
    by_lot[np.flatnonzero(priced_mask)] = priced

    return ElementResult(
        name=element.name,
        model=model,
        projections=projections,
        summary=summary,
        ctx=ctx,
        total=float(by_lot.sum()),
        by_lot=by_lot,
        n_lots_fitted=int(ctx["n_keep"]),
        kind="fitted",
    )


def _selected_model(summary: pd.DataFrame) -> str:
    """Which model the engine picked, without needing cost_core installed."""
    row = summary.loc[summary["Item"] == "SELECTED"]
    if len(row):
        for name in ("LC", "Rate", "LC+Rate"):
            if str(row.iloc[0][name]).strip().upper() == "YES":
                return name
    raise ProgramError("No model was selected; nothing could be fitted.")


#: How far the fitted spec may sit from the element's own simulated
#: percentiles before it stops being a fair summary of it. Measured, not
#: guessed: see tests/test_wbs.py::TestDistributionHandoff.
SPEC_TOLERANCE = 0.015


def _lognormal_spec(totals: np.ndarray) -> dict:
    """Describe an element's simulated total as a lognormal.

    cost_core's WBS model takes distributions rather than raw draws, so each
    element's simulated total has to be summarised by a two-parameter family
    before it can be correlated with the others. That step costs accuracy: an
    element total is a sum of six-odd correlated lognormal lot costs, which is
    not itself lognormal, so no two-parameter fit reproduces it exactly.

    Matching in log space was chosen by measurement rather than by taste. It
    tracks the element's own percentiles to about half a percent, and beat
    both arithmetic moment-matching and a normal on the same data. The median
    comes back within 0.05%; the upper tail is the part that gives, sitting
    around half a percent high at P80. A test bounds it at
    :data:`SPEC_TOLERANCE` so it cannot quietly get worse.
    """
    logs = np.log(np.clip(totals, 1e-12, None))
    return {
        "type": "lognormal",
        "mean": float(np.mean(logs)),
        "sigma": float(max(np.std(logs, ddof=1), 1e-9)),
    }


def roll_up(
    program: Program,
    overrides: dict | None = None,
    *,
    simulate: bool = True,
    correlation: float = DEFAULT_CORRELATION,
    n_iter: int = 20000,
    seed: int = 11,
) -> ProgramResult:
    """Price every element and add them into one programme estimate.

    Raises:
        ProgramError: If the programme is malformed, or an element cannot be
            fitted. One element failing stops the roll-up rather than
            producing a total that quietly omits it.
    """
    program.validate()

    priced: dict[str, ElementResult] = {}
    for el in program.pricing_order():
        if el.kind == "fitted":
            priced[el.name] = price_element(program, el, overrides)
        else:
            priced[el.name] = price_derived(program, el, priced)
    # Report in the order the analyst entered them, not the order forced by
    # the dependencies between them.
    results = [priced[el.name] for el in program.elements]

    by_lot = pd.DataFrame({"Lot": range(1, program.n_lots + 1),
                           "Fiscal Year": program.fiscal_years})
    for r in results:
        by_lot[r.name] = np.round(r.by_lot, 2)
    by_lot["Program Total ($)"] = np.round(
        np.sum([r.by_lot for r in results], axis=0), 2
    )

    notes = [
        f"{len(results)} WBS element(s), each fitted to its own analogy "
        "history and priced against the shared lot schedule.",
        "Element quantities differ where the buy differs, so a kit or a "
        "spares provision changes the count without changing the schedule.",
    ]
    warnings: list[str] = []

    result = ProgramResult(
        program=program.name,
        order=program.pricing_order(),
        elements=results,
        by_lot=by_lot,
        total=float(by_lot["Program Total ($)"].sum()),
        notes=notes,
        warnings=warnings,
    )

    if simulate:
        if not RISK_AVAILABLE:
            warnings.append(
                "cost_core is not installed, so the programme total carries "
                f"no risk analysis. {RISK_IMPORT_ERROR}"
            )
        else:
            _program_risk(result, correlation, n_iter, seed)

    return result


def _program_risk(
    result: ProgramResult, correlation: float, n_iter: int, seed: int
):
    """Simulate each fitted element from its own history, then correlate.

    Only fitted elements carry uncertainty measured from data, so only they
    go into the copula. The other two kinds are derived from that same draw
    rather than given a spread of their own:

    A factor element is a fixed percentage of other elements, so on every
    iteration it is exactly that percentage of whatever they came out at. It
    is perfectly correlated with its basis by construction, which is both
    true and stronger than any correlation the model could be told to assume.

    An amount element is a number the analyst entered. Inventing a
    distribution around it would be inventing uncertainty nobody measured, so
    it is carried as fixed and contributes no variance. Its cost is still in
    every percentile, it just does not move.
    """
    by_name = {r.name: r for r in result.elements}
    fitted = [r for r in result.elements if r.kind == "fitted"]

    elements = []
    for r in fitted:
        buy = simulate_buy(
            r.ctx,
            r.projections,
            r.model,
            n_iter=n_iter,
            seed=seed,
            lot_correlation=0.30,
        )
        r.totals = np.asarray(buy.totals, dtype=float)
        elements.append(
            CostElement(
                name=r.name,
                distribution=_lognormal_spec(r.totals),
                point_estimate=r.total,
            )
        )

    model = RiskModel(
        elements=elements,
        default_correlation=float(correlation),
        name=result.program,
    )

    import warnings as _warnings

    with _warnings.catch_warnings(record=True) as caught:
        _warnings.simplefilter("always")
        sim = simulate_risk_model(model, n_iter=n_iter, seed=seed)
    for w in caught:
        text = str(w.message)
        if text not in result.warnings:
            result.warnings.append(text)

    # Correlated draws per fitted element, which the derived kinds build on.
    samples = np.asarray(sim.element_samples, dtype=float)
    names = list(sim.element_names)
    draws = {name: samples[:, i] for i, name in enumerate(names)}

    for r in fitted:
        if r.name in draws:
            r.totals = draws[r.name]

    # Derived elements, in the order their bases were priced.
    for el in result.order:
        r = by_name[el.name]
        if r.kind == "amount":
            r.totals = np.full(samples.shape[0], r.total, dtype=float)
            draws[r.name] = r.totals
        elif r.kind == "factor":
            base = np.sum([draws[b] for b in r.basis], axis=0)
            share = r.total / float(np.sum([by_name[b].total for b in r.basis]))                 if np.sum([by_name[b].total for b in r.basis]) else 0.0
            r.totals = share * base
            draws[r.name] = r.totals

    totals = np.sum([draws[r.name] for r in result.elements], axis=0)

    result.p50 = float(np.percentile(totals, 50))
    result.p80 = float(np.percentile(totals, 80))
    result.p90 = float(np.percentile(totals, 90))
    result.mean = float(np.mean(totals))
    result.std = float(np.std(totals, ddof=1))
    result.cv = float(result.std / result.mean) if result.mean else float("nan")
    result.point_percentile = float((totals < result.total).mean() * 100.0)
    result.correlation = float(correlation)
    result.n_iter = int(n_iter)
    result.scurve = _scurve(totals)
    result.tornado = _tornado(result, draws, totals)

    if len(elements) > 1:
        try:
            impact = correlation_impact(model, n_iter=n_iter, seed=seed)
            result.independence_understates_sd_by = float(
                np.sqrt(impact.empirical_variance_ratio)
            )
            result.variance_ratio_analytic = float(
                impact.analytic_variance_ratio
            )
            result.p80_understatement = float(impact.p80_understatement)
            result.reserve_understatement = float(
                impact.reserve_understatement
            )
        except Exception as exc:
            result.warnings.append(
                f"Could not measure the cost of assuming independence: {exc}"
            )

    result.notes.append(
        f"Element totals are correlated at {correlation:.2f}. Treating them "
        "as independent would let their overruns cancel and understate the "
        "spread of the programme total."
    )
    result.notes.append(
        "Each fitted element is summarised as a lognormal before being "
        "correlated with the others, which tracks its own percentiles to "
        "about half a percent. Read the programme percentiles at that "
        "resolution."
    )
    derived = [r for r in result.elements if r.kind != "fitted"]
    if derived:
        result.notes.append(
            "Factor elements move exactly with what they are a percentage "
            "of, and amount elements are carried as entered. Neither is "
            "given a spread of its own, because neither has a history to "
            "measure one from."
        )


def _tornado(result: ProgramResult, draws: dict, totals: np.ndarray):
    """Variance contribution by element, over every kind.

    The same decomposition cost_core uses, Cov(X_i, T)/Var(T), applied to the
    full set rather than the fitted elements alone. It still sums to one,
    because the total is the sum of its parts. An amount element lands at
    exactly zero, which is the honest answer for a number that does not move.
    """
    var_total = float(np.var(totals, ddof=1))
    rows = []
    for r in result.elements:
        x = draws.get(r.name)
        if x is None:
            continue
        cov = float(np.cov(x, totals, ddof=1)[0, 1])
        rows.append(
            {
                "component": r.name,
                "kind": r.kind,
                "std_dev": float(np.std(x, ddof=1)),
                "covariance_with_total": cov,
                "variance_share": cov / var_total if var_total else 0.0,
            }
        )
    frame = pd.DataFrame(rows)
    if len(frame):
        frame = frame.sort_values(
            "variance_share", ascending=False
        ).reset_index(drop=True)
    return frame


def _scurve(totals: np.ndarray, step: int = 1) -> pd.DataFrame:
    pct = np.arange(step, 100, step)
    return pd.DataFrame(
        {
            "Percentile": pct / 100.0,
            "Program Total ($)": np.round(np.percentile(totals, pct), 2),
        }
    )


def element_summary(result: ProgramResult) -> pd.DataFrame:
    """One row per element, for the roll-up sheet and the results pane.

    Only fitted elements have a T1 or a unit count, so the other kinds leave
    those blank rather than showing a zero that reads like a real number.
    """
    rows = []
    for r in result.elements:
        share = r.total / result.total if result.total else float("nan")
        row = {
            "WBS Element": r.name,
            "Kind": r.kind,
            "Model": r.model,
            "Analogy lots": r.n_lots_fitted if r.kind == "fitted" else "",
            "T1 ($K)": "",
            "Units bought": "",
            "Cost Before Risk ($)": round(r.total, 2),
            "Share of Program": round(share, 4),
        }
        if r.kind == "fitted":
            col = f"{r.model} T1 First Unit Cost ($K)"
            if col in r.projections.columns and len(r.projections):
                row["T1 ($K)"] = r.projections[col].iloc[0]
            if "Lot Quantity" in r.projections.columns:
                row["Units bought"] = int(r.projections["Lot Quantity"].sum())
        if r.totals is not None:
            row["P80 With Risk ($)"] = round(
                float(np.percentile(r.totals, 80)), 2
            )
        rows.append(row)

    total_row = {
        "WBS Element": "PROGRAM TOTAL",
        "Kind": "",
        "Model": "",
        "Analogy lots": "",
        "T1 ($K)": "",
        "Units bought": "",
        "Cost Before Risk ($)": round(result.total, 2),
        "Share of Program": 1.0,
    }
    if result.p80 is not None:
        # Not the sum of the element P80s, and deliberately so: they do not
        # all peak together. See the correlated roll-up.
        total_row["P80 With Risk ($)"] = round(result.p80, 2)
    rows.append(total_row)
    return pd.DataFrame(rows)


def program_summary(result: ProgramResult) -> pd.DataFrame:
    """The programme headline, as a two-column table."""
    def money(v):
        return "n/a" if v is None else f"{v:,.2f}"

    rows = [
        ("Program", result.program),
        ("WBS elements", str(len(result.elements))),
        ("Lots", str(len(result.by_lot))),
        ("", ""),
        ("--- BEFORE RISK ---", "deterministic estimate, no risk applied"),
        ("Program total before risk ($)", money(result.total)),
    ]
    for r in result.elements:
        rows.append((f"    {r.name} ({r.model})", money(r.total)))

    if result.p50 is None:
        rows += [
            ("", ""),
            ("Risk analysis", "not run"),
        ]
    else:
        rows += [
            ("", ""),
            ("--- WITH RISK ---", "correlated Monte Carlo over the elements"),
            ("Element correlation", f"{result.correlation:.2f}"),
            ("Monte Carlo iterations", f"{result.n_iter:,}"),
            ("Program P50 ($)", money(result.p50)),
            ("Program P80 ($)", money(result.p80)),
            ("Program P90 ($)", money(result.p90)),
            ("Program CV", f"{result.cv:.4f}"),
            (
                "Reserve to P80 ($)",
                money(result.p80 - result.total)
                + "  (the amount risk adds to the estimate above)",
            ),
            (
                "Point estimate falls at",
                f"P{result.point_percentile:.0f} of the simulated program",
            ),
        ]
        if result.independence_understates_sd_by:
            rows.append((
                "Independence would understate spread by",
                f"{result.independence_understates_sd_by:.2f}x on the "
                "standard deviation "
                f"(variance ratio {result.variance_ratio_analytic:.2f} "
                "by the closed form)",
            ))
        if result.p80_understatement is not None:
            rows.append((
                "P80 under independence would be low by",
                f"{result.p80_understatement:.2%}",
            ))
        if result.reserve_understatement is not None:
            rows.append((
                "Share of the P80 reserve independence throws away",
                f"{result.reserve_understatement:.1%}",
            ))
    for i, n in enumerate(result.notes):
        rows.append(("Note" if i == 0 else "", n))
    for i, w in enumerate(result.warnings):
        rows.append(("Warning" if i == 0 else "", w))
    return pd.DataFrame(rows, columns=["Item", "Value"])


def save_program_workbook(
    filename: str,
    result: ProgramResult,
    sensitivity: pd.DataFrame | None = None,
):
    """Write the roll-up: programme first, then every element behind it.

    The order is deliberate. The programme total is what gets briefed, but a
    reviewer's first question is which element it came from, so the element
    breakdown sits immediately behind the headline rather than at the end.
    """
    import openpyxl
    from openpyxl.chart import Reference, ScatterChart, Series

    from lot_cost_model import (
        _format_chart,
        _money_axis_fmt,
        _money_short,
        _nice_bounds,
    )

    element_table = element_summary(result)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        program_summary(result).to_excel(
            writer, sheet_name="Program_Summary", index=False
        )
        element_table.to_excel(
            writer, sheet_name="Program_Elements", index=False
        )
        result.by_lot.to_excel(
            writer, sheet_name="Program_By_Lot", index=False
        )
        if result.scurve is not None:
            result.scurve.to_excel(
                writer, sheet_name="Program_SCurve", index=False
            )
        if result.tornado is not None and len(result.tornado):
            result.tornado.to_excel(
                writer, sheet_name="Program_Tornado", index=False
            )
        if sensitivity is not None and len(sensitivity):
            sensitivity.to_excel(
                writer, sheet_name="Buy_Sensitivity", index=False
            )
        infl = influence_table(result)
        if infl is not None:
            infl.to_excel(
                writer, sheet_name="Element_Influence", index=False
            )
        used = {
            "Program_Summary", "Program_Elements", "Program_By_Lot",
            "Program_SCurve", "Program_Tornado", "Buy_Sensitivity",
            "Element_Influence",
        }
        for r in result.elements:
            r.projections.to_excel(
                writer, sheet_name=_sheet_name(r.name, used), index=False
            )

    wb = openpyxl.load_workbook(filename)

    # Cost by fiscal year, stacked so the element mix is visible.
    ws = wb["Program_By_Lot"]
    last = len(result.by_lot) + 1
    if last > 1:
        from openpyxl.chart import BarChart

        bar = BarChart()
        bar.type = "col"
        bar.grouping = "stacked"
        bar.overlap = 100
        _format_chart(
            bar, f"{result.program}: cost by fiscal year",
            "Fiscal Year", "Cost ($)", 20, 11,
        )
        names = [r.name for r in result.elements]
        for i, _ in enumerate(names):
            bar.series.append(
                Series(
                    Reference(ws, min_col=3 + i, min_row=1, max_row=last),
                    title_from_data=True,
                )
            )
        bar.set_categories(
            Reference(ws, min_col=2, min_row=2, max_row=last)
        )
        # A fiscal year is a label, not a quantity: the shared chart format
        # would otherwise render 2028 as "2,028".
        bar.x_axis.numFmt = "0"
        bar.y_axis.numFmt = '#,##0,,"M"' 
        ws.add_chart(bar, f"{_col_letter(len(names) + 5)}2")

    if result.tornado is not None and len(result.tornado):
        from openpyxl.chart import BarChart

        wst = wb["Program_Tornado"]
        cols = list(result.tornado.columns)
        share_col = cols.index("variance_share") + 1
        last_t = len(result.tornado) + 1
        tor = BarChart()
        tor.type = "bar"          # horizontal, which is what makes it a tornado
        # On a horizontal bar chart openpyxl's x_axis is still the category
        # axis, drawn down the side, and y_axis is the value axis along the
        # bottom. Naming them the other way round leaves the shares reading
        # as "0" under the shared numeric format.
        _format_chart(
            tor, f"{result.program}: share of program variance",
            "WBS element", "Share of variance", 18, 10,
        )
        tor.y_axis.numFmt = "0%"
        tor.legend = None
        tor.series.append(
            Series(
                Reference(wst, min_col=share_col, min_row=1, max_row=last_t),
                title_from_data=True,
            )
        )
        tor.set_categories(
            Reference(wst, min_col=1, min_row=2, max_row=last_t)
        )
        wst.add_chart(tor, f"{_col_letter(len(cols) + 2)}2")

    if sensitivity is not None and len(sensitivity):
        wss = wb["Buy_Sensitivity"]
        cols = list(sensitivity.columns)
        last_s = len(sensitivity) + 1
        unit_col = cols.index("Cost per Unit ($)") + 1
        chart = ScatterChart()
        _format_chart(
            chart, f"{result.program}: unit cost against buy size",
            "Buy multiplier", "Cost per unit ($)", 18, 10,
        )
        chart.x_axis.numFmt = "0.0"
        chart.legend = None
        lo, hi = _nice_bounds(
            float(sensitivity["Cost per Unit ($)"].min()),
            float(sensitivity["Cost per Unit ($)"].max()),
        )
        chart.y_axis.numFmt = _money_axis_fmt(lo, hi)
        chart.y_axis.scaling.min = max(0.0, lo)
        chart.y_axis.scaling.max = hi
        s_unit = Series(
            values=Reference(wss, min_col=unit_col, min_row=1, max_row=last_s),
            xvalues=Reference(wss, min_col=1, min_row=2, max_row=last_s),
            title_from_data=True,
        )
        s_unit.marker.symbol = "circle"
        s_unit.marker.size = 8
        s_unit.smooth = False
        chart.series.append(s_unit)
        wss.add_chart(chart, f"{_col_letter(len(cols) + 2)}2")

    if result.scurve is not None and len(result.scurve):
        _add_program_scurve(wb, result, _format_chart, _money_short,
                            _nice_bounds, _money_axis_fmt, ScatterChart,
                            Series, Reference)

    wb.save(filename)


def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(idx)


def _sheet_name(name: str, taken: set[str] | None = None) -> str:
    """A legal, unique Excel sheet name for this element.

    Excel refuses several characters and truncates at 31, which is short
    enough that two real WBS names can collide: "1.1 Air Vehicle Structural
    Assembly Group A" and "... Group B" share their first 31 characters. Left
    alone that either errors or drops an element's sheet, so a numeric suffix
    is added when the truncated name is already spoken for.
    """
    cleaned = "".join(c for c in name if c not in r"[]:*?/\\").strip()
    base = (cleaned[:31] or "Element").strip()
    if taken is None or base not in taken:
        if taken is not None:
            taken.add(base)
        return base
    for n in range(2, 100):
        suffix = f" ({n})"
        candidate = (cleaned[: 31 - len(suffix)]).strip() + suffix
        if candidate not in taken:
            taken.add(candidate)
            return candidate
    raise ProgramError(
        f"Cannot make a unique sheet name for {name!r}; shorten the element "
        "names so their first 31 characters differ."
    )


def _add_program_scurve(
    wb, result, _format_chart, _money_short, _nice_bounds,
    _money_axis_fmt, ScatterChart, Series, Reference,
):
    ws = wb["Program_SCurve"]
    last = len(result.scurve) + 1
    for row in range(2, last + 1):
        ws.cell(row=row, column=1).number_format = "0%"
        ws.cell(row=row, column=2).number_format = "#,##0"

    curve = ScatterChart()
    _format_chart(
        curve,
        f"{result.program}: probability the program comes in at or below",
        "Program Total ($)", "Cumulative Probability", 20, 11,
    )
    curve.y_axis.numFmt = "0%"
    curve.y_axis.scaling.min = 0
    curve.y_axis.scaling.max = 1

    lo = float(result.scurve["Program Total ($)"].min())
    hi = float(result.scurve["Program Total ($)"].max())
    curve.x_axis.numFmt = _money_axis_fmt(lo, hi)
    nice_lo, nice_hi = _nice_bounds(lo, hi)
    curve.x_axis.scaling.min = max(0.0, nice_lo)
    curve.x_axis.scaling.max = nice_hi

    from openpyxl.chart.series import SeriesLabel

    line = Series(
        values=Reference(ws, min_col=1, min_row=1, max_row=last),
        xvalues=Reference(ws, min_col=2, min_row=2, max_row=last),
        title_from_data=True,
    )
    line.marker.symbol = "none"
    line.smooth = True
    line.tx = SeriesLabel(v="Program total distribution")
    curve.series.append(line)

    from openpyxl.chart.shapes import GraphicalProperties

    pct = result.scurve["Percentile"].round(4)
    for i, (label, level, colour, symbol) in enumerate(
        [("P50", 0.50, "0072B2", "circle"), ("P80", 0.80, "D55E00", "diamond")]
    ):
        hit = result.scurve.loc[pct == round(level, 4), "Program Total ($)"]
        if hit.empty:
            continue
        cost = float(hit.iloc[0])
        y_col, x_col = 4 + i * 2, 5 + i * 2
        ws.cell(row=1, column=y_col, value=f"{label}  {_money_short(cost)}")
        ws.cell(row=2, column=y_col, value=level).number_format = "0%"
        ws.cell(row=1, column=x_col, value=f"{label} cost")
        ws.cell(row=2, column=x_col, value=cost).number_format = "#,##0"

        mark = Series(
            values=Reference(ws, min_col=y_col, min_row=1, max_row=2),
            xvalues=Reference(ws, min_col=x_col, min_row=2, max_row=2),
            title_from_data=True,
        )
        mark.marker.symbol = symbol
        mark.marker.size = 11
        style = GraphicalProperties(solidFill=colour)
        style.line.solidFill = "FFFFFF"
        style.line.width = 19050
        mark.marker.graphicalProperties = style
        mark.graphicalProperties.line.noFill = True
        curve.series.append(mark)

    ws.add_chart(curve, "I2")


def influence(result: ElementResult) -> pd.DataFrame | None:
    """Which analogy lot is actually carrying this element's fit.

    Six analogy lots is normal, and at that size one lot can set the slope
    while every summary statistic still looks healthy. Leverage says which lot
    is unusual among the predictors; Cook's distance says which one is moving
    the fit. Both are flags rather than verdicts: the largest or smallest lot
    in a sample has high leverage by construction, and dropping it for that
    reason alone would be indefensible.
    """
    if not RISK_AVAILABLE:
        return None
    try:
        return influence_diagnostics(result.ctx, result.model)
    except Exception:
        return None


def _scale_element(el: Element, factor: float) -> Element:
    """A copy of this element priced for a buy of ``factor`` times the size.

    What scales depends on the kind, and getting that wrong is the whole
    point of having kinds:

    A fitted element's quantities scale, rounded to whole units, because you
    cannot buy 7.2 airframes. A lot it sits out stays at zero.

    A factor element is left alone. It is a percentage of its basis, so it
    follows the hardware down or up without being touched.

    An amount element is also left alone, and deliberately. Tooling and
    qualification are nonrecurring: buying forty percent fewer aircraft does
    not buy forty percent less tooling. Scaling them would quietly turn a
    one-off into a variable cost and flatter every small-buy case.
    """
    if el.kind != "fitted":
        return replace(el)
    return replace(
        el,
        quantities=[
            0.0 if q <= 0 else float(max(1, round(q * factor)))
            for q in el.quantities
        ],
    )


def buy_profile_sensitivity(
    program: Program,
    factors: "list[float] | tuple[float, ...]" = (0.6, 0.8, 1.0, 1.2, 1.5),
    overrides: dict | None = None,
    reference_element: str | None = None,
) -> pd.DataFrame:
    """Reprice the whole programme at several buy sizes.

    Every element's lot quantities scale together, so a two-per-aircraft
    engine count and a spares provision keep their proportion to the end item
    rather than drifting. Quantities are whole units, and a lot an element
    sits out stays at zero.

    This is the question the rate term exists to answer, and the reason unit
    cost moves at all: buying fewer units per lot pushes the cost of each one
    up, on top of the learning that is lost by building fewer of them.

    Args:
        program: The programme to reprice. It is not modified.
        factors: Multipliers on the baseline buy.
        overrides: Settings passed through to the engine.
        reference_element: Which element counts as the end item for the
            per-unit column. Defaults to the first.

    Raises:
        ProgramError: If a factor prices out to nothing, or the reference
            element is not in the programme.
    """
    program.validate()
    names = [e.name for e in program.elements]
    reference = reference_element or names[0]
    if reference not in names:
        raise ProgramError(
            f"{reference!r} is not an element of this programme. Pick one of "
            f"{names}."
        )

    baseline_total = None
    rows = []
    for factor in factors:
        scaled = Program(
            name=program.name,
            fiscal_years=list(program.fiscal_years),
            elements=[_scale_element(el, factor) for el in program.elements],
        )
        try:
            priced = roll_up(scaled, overrides, simulate=False)
        except ProgramError as exc:
            raise ProgramError(f"At {factor:g}x the buy: {exc}") from exc

        end_items = int(
            sum(
                e.projections["Lot Quantity"].sum()
                for e in priced.elements
                if e.name == reference
            )
        )
        if baseline_total is None and factor == 1.0:
            baseline_total = priced.total

        rows.append(
            {
                "Buy Multiplier": factor,
                f"{reference} Units": end_items,
                "Program Total ($)": round(priced.total, 2),
                "Cost per Unit ($)": round(priced.total / end_items, 2)
                if end_items
                else float("nan"),
            }
        )

    frame = pd.DataFrame(rows)
    if baseline_total:
        frame["Change vs Baseline"] = (
            frame["Program Total ($)"] / baseline_total - 1.0
        ).round(4)
        base_unit = frame.loc[
            frame["Buy Multiplier"] == 1.0, "Cost per Unit ($)"
        ]
        if len(base_unit):
            frame["Unit Cost vs Baseline"] = (
                frame["Cost per Unit ($)"] / float(base_unit.iloc[0]) - 1.0
            ).round(4)
    return frame


def influence_table(result: ProgramResult) -> pd.DataFrame | None:
    """Every element's influence diagnostics in one table."""
    frames = []
    for r in result.elements:
        one = influence(r)
        if one is None or one.empty:
            continue
        one = one.copy()
        one.insert(0, "WBS Element", r.name)
        frames.append(one)
    if not frames:
        return None
    return pd.concat(frames, ignore_index=True)
